"""Парсинг «Ядро.xlsx»: листы Подрядчики и Объекты."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_EMPTY_MARKERS = frozenset({"", "—", "-", "–", "None"})


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: str) -> bool:
    return _cell_str(value) in _EMPTY_MARKERS


def parse_core_contractors(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Подрядчики" not in wb.sheetnames:
        wb.close()
        raise ValueError('В файле нет листа «Подрядчики»')
    ws = wb["Подрядчики"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Лист «Подрядчики» пуст")
    headers = [_cell_str(h) for h in rows[0]]
    expected = {
        "Файл",
        "Инспекция проведена",
        "Генподрядчик",
        "Субподрядчик",
        "Договор",
        "Контактное лицо",
        "Тел.",
        "Факс",
        "E-Mail",
        "Доп. информация",
        "Примечание (расхождения)",
    }
    if not expected.issubset(set(headers)):
        raise ValueError(f"Неожиданные заголовки на листе «Подрядчики»: {headers}")

    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        file_label = _cell_str(rec.get("Файл"))
        gen = _cell_str(rec.get("Генподрядчик"))
        if not file_label and not gen:
            continue
        out.append(
            {
                "file_label": file_label,
                "inspection_type": _cell_str(rec.get("Инспекция проведена")),
                "gen_contractor": gen,
                "sub_contractor": _cell_str(rec.get("Субподрядчик")),
                "contract_no": _cell_str(rec.get("Договор")),
                "contact_person": _cell_str(rec.get("Контактное лицо")),
                "contact_phone": _cell_str(rec.get("Тел.")),
                "contact_fax": _cell_str(rec.get("Факс")),
                "contact_email": _cell_str(rec.get("E-Mail")),
                "extra_info": _cell_str(rec.get("Доп. информация")),
                "note_discrepancy": _cell_str(rec.get("Примечание (расхождения)")),
            }
        )
    if not out:
        raise ValueError("На листе «Подрядчики» нет данных")
    return out


def parse_core_objects(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Объекты" not in wb.sheetnames:
        wb.close()
        raise ValueError('В файле нет листа «Объекты»')
    ws = wb["Объекты"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Лист «Объекты» пуст")
    headers = [_cell_str(h) for h in rows[0]]
    if headers[:4] != ["№", "Организация", "Шифр", "Объект"]:
        raise ValueError(f"Неожиданные заголовки на листе «Объекты»: {headers[:4]}")

    out: list[dict[str, Any]] = []
    org: str = ""
    for row in rows[1:]:
        if not any(row):
            continue
        org_cell = _cell_str(row[1] if len(row) > 1 else "")
        if org_cell:
            org = org_cell
        cipher = _cell_str(row[2] if len(row) > 2 else "")
        object_name = _cell_str(row[3] if len(row) > 3 else "")
        if not cipher and not object_name:
            continue
        if not org:
            raise ValueError(f"Объект без организации: шифр={cipher!r}")
        if not cipher:
            raise ValueError(f"Строка без шифра (организация {org!r})")
        out.append(
            {
                "org_label": org,
                "cipher": cipher,
                "object_name": object_name or cipher,
            }
        )
    if not out:
        raise ValueError("На листе «Объекты» нет данных")
    return out
