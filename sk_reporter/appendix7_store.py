"""Приложение 7 — заполнение xlsm из luvr.yaml (макросы в шаблоне не трогаем)."""

from __future__ import annotations

import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from sk_reporter.luvr_store import _norm_fio
from sk_reporter.paths import luvr_dir, repo_root

_A7_SHEET = "Отчет о ВОУ"
_MONTH_GENITIVE = {
    "Январь": "января",
    "Февраль": "февраля",
    "Март": "марта",
    "Апрель": "апреля",
    "Май": "мая",
    "Июнь": "июня",
    "Июль": "июля",
    "Август": "августа",
    "Сентябрь": "сентября",
    "Октябрь": "октября",
    "Ноябрь": "ноября",
    "Декабрь": "декабря",
}


def load_luvr_manifest() -> dict[str, Any]:
    path = luvr_dir() / "manifest.yaml"
    if not path.is_file():
        return {}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def appendix7_template_path() -> Path:
    manifest = load_luvr_manifest()
    name = manifest.get("appendix7")
    if not name:
        raise FileNotFoundError("В manifest.yaml не задан appendix7")
    path = luvr_dir() / str(name)
    if not path.is_file():
        raise FileNotFoundError(f"Шаблон Прил.7 не найден: {path.name}")
    return path


def appendix7_output_dir() -> Path:
    out = luvr_dir() / "generated"
    out.mkdir(parents=True, exist_ok=True)
    return out


def _mark_to_a7_value(mark: str) -> float | int:
    if not mark:
        return 0
    if mark == "1":
        return 1
    if mark == "0.5":
        return 0.5
    try:
        return float(str(mark).replace(",", "."))
    except ValueError:
        return 0


def _parse_iso_date(v: str | None) -> date | None:
    if not v:
        return None
    try:
        return date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _find_day_columns_ws(ws) -> tuple[int, list[dict[str, Any]]]:
    """Строка с датами — та, где больше всего ячеек datetime (обычно 12-я)."""
    best_row = 0
    best_cols: list[dict[str, Any]] = []
    max_r = min(30, ws.max_row or 0)
    max_c = ws.max_column or 0
    for r in range(1, max_r + 1):
        cols: list[dict[str, Any]] = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                cols.append({"col": c, "date": v.date()})
            elif isinstance(v, date):
                cols.append({"col": c, "date": v})
        if len(cols) > len(best_cols):
            best_row = r
            best_cols = cols
    if len(best_cols) < 7:
        return 0, []
    return best_row, best_cols


def _is_summary_fio_row(ws, row: int) -> str | None:
    c1 = ws.cell(row=row, column=1).value
    c2 = ws.cell(row=row, column=2).value
    c3 = ws.cell(row=row, column=3).value
    c4 = ws.cell(row=row, column=4).value
    c5 = ws.cell(row=row, column=5).value
    if not c3 or not isinstance(c3, str):
        return None
    parts = _norm_fio(c3).split()
    if len(parts) < 3:
        return None
    if c1 or c2 or c4 or c5:
        return None
    return _norm_fio(c3)


def _period_line(month_sheet: str, days: list[dict[str, Any]]) -> str | None:
    if not days:
        return None
    gen = _MONTH_GENITIVE.get(month_sheet)
    first = _parse_iso_date(days[0].get("date"))
    last = _parse_iso_date(days[-1].get("date"))
    if not gen or not first or not last:
        return None
    return (
        f"за отчетный месяц (период) с {first.day:02d} {gen} {first.year} "
        f"по {last.day:02d} {gen} {last.year}"
    )


def appendix7_status() -> dict[str, Any]:
    folder = luvr_dir()
    manifest = load_luvr_manifest()
    template = None
    template_error = None
    try:
        template = appendix7_template_path()
    except FileNotFoundError as e:
        template_error = str(e)

    summary_rows = 0
    day_count = 0
    if template is not None:
        wb = load_workbook(template, read_only=True, data_only=True, keep_vba=True)
        if _A7_SHEET in wb.sheetnames:
            ws = wb[_A7_SHEET]
            _, day_cols = _find_day_columns_ws(ws)
            day_count = len(day_cols)
            for r in range(1, (ws.max_row or 0) + 1):
                if _is_summary_fio_row(ws, r):
                    summary_rows += 1
        wb.close()

    return {
        "template_present": template is not None,
        "template_name": template.name if template else manifest.get("appendix7"),
        "template_error": template_error,
        "sheet_name": _A7_SHEET,
        "summary_rows": summary_rows,
        "day_columns": day_count,
        "output_dir": str(appendix7_output_dir().relative_to(repo_root())),
    }


def build_appendix7_from_luvr(month_sheet: str) -> dict[str, Any]:
    from sk_reporter.luvr_store import luvr_month_payload

    month = luvr_month_payload(month_sheet)
    people = month.get("people") or []
    days = month.get("days") or []
    if not people or not days:
        raise ValueError(f"Нет данных ЛУВР для «{month_sheet}»")

    template = appendix7_template_path()
    year = month.get("year") or (_parse_iso_date(days[0].get("date")) or date.today()).year
    safe_sheet = re.sub(r"[^\w\-]+", "_", month_sheet)
    output = appendix7_output_dir() / f"Прил7_{safe_sheet}_{year}.xlsm"
    shutil.copy2(template, output)

    luvr_by_fio: dict[str, dict[str, Any]] = {}
    for person in people:
        key = _norm_fio(person.get("fio", "")).lower()
        if key:
            luvr_by_fio[key] = person

    wb = load_workbook(output, keep_vba=True)
    if _A7_SHEET not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Лист «{_A7_SHEET}» не найден в шаблоне Прил.7")

    ws = wb[_A7_SHEET]
    day_row, day_cols = _find_day_columns_ws(ws)
    if not day_cols:
        wb.close()
        raise ValueError("В шаблоне не найдена строка с датами")

    period = _period_line(month_sheet, days)
    if period:
        ws.cell(row=8, column=2).value = period

    for i, dc in enumerate(day_cols):
        if i >= len(days):
            break
        d = _parse_iso_date(days[i].get("date"))
        if d:
            ws.cell(row=day_row, column=dc["col"]).value = datetime(d.year, d.month, d.day)

    filled_rows = 0
    cells_updated = 0
    matched_fio: set[str] = set()
    template_fio: set[str] = set()

    for r in range(day_row + 1, (ws.max_row or 0) + 1):
        fio = _is_summary_fio_row(ws, r)
        if not fio:
            continue
        template_fio.add(fio.lower())
        person = luvr_by_fio.get(fio.lower())
        if not person:
            continue
        marks = person.get("marks") or []
        row_updates = 0
        for i, dc in enumerate(day_cols):
            if i >= len(marks):
                break
            val = _mark_to_a7_value(marks[i])
            cell = ws.cell(row=r, column=dc["col"])
            if cell.value != val:
                cell.value = val
                row_updates += 1
        if row_updates:
            filled_rows += 1
            cells_updated += row_updates
            matched_fio.add(fio.lower())

    for ws_sheet in wb.worksheets:
        ws_sheet.legacy_drawing = None

    wb.save(output)
    wb.close()

    unmatched_luvr = sorted(
        _norm_fio(p.get("fio", ""))
        for p in people
        if _norm_fio(p.get("fio", "")).lower() not in template_fio
    )
    unmatched_template = sorted(
        _norm_fio(fio)
        for fio in template_fio
        if fio not in matched_fio
    )

    return {
        "ok": True,
        "sheet": month_sheet,
        "year": year,
        "output": str(output),
        "output_name": output.name,
        "download": f"/download/luvr/generated/{output.name}",
        "filled_rows": filled_rows,
        "cells_updated": cells_updated,
        "matched_people": len(matched_fio),
        "unmatched_luvr": unmatched_luvr[:20],
        "unmatched_luvr_count": len(unmatched_luvr),
        "unmatched_template_count": len(unmatched_template),
        "note": "Заполнены только сводные строки по ФИО; разбивка по объектам — в Excel вручную/макросами.",
    }
