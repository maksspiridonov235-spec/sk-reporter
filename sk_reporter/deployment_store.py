"""Расстановка — заполнение xlsm из project.yaml + personnel (или из ЛУВР)."""

from __future__ import annotations

import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sk_reporter.appendix7_store import appendix7_output_dir, load_luvr_manifest
from sk_reporter.luvr_store import _norm_fio
from sk_reporter.paths import luvr_dir, repo_root

_DEPLOY_SHEET = "Отчет расстановка"
_FIO_SHEET = "Список ФИО"
_POS_SHEET = "Персонал Свод"
_DATA_START_ROW = 4
_DEFAULT_CONTROL = "Инспекционный контроль, Проверка ИТД"


def deployment_template_path() -> Path:
    manifest = load_luvr_manifest()
    name = manifest.get("deployment")
    if not name:
        raise FileNotFoundError("В manifest.yaml не задан deployment")
    path = luvr_dir() / str(name)
    if not path.is_file():
        raise FileNotFoundError(f"Шаблон расстановки не найден: {path.name}")
    return path


def deployment_output_dir() -> Path:
    return appendix7_output_dir()


def _phone_digits(phone: str) -> int | str:
    digits = re.sub(r"\D", "", phone or "")
    if not digits:
        return ""
    try:
        return int(digits)
    except ValueError:
        return digits


def _deployment_rows_from_projects() -> tuple[list[dict[str, Any]], dict[str, int]]:
    from sk_reporter.personnel_store import get_person
    from sk_reporter.project_store import list_projects_rich

    rows: list[dict[str, Any]] = []
    stats = {"projects_total": 0, "projects_with_engineers": 0, "projects_empty": 0}

    for proj in list_projects_rich():
        stats["projects_total"] += 1
        engineer_ids = proj.get("engineer_ids") or []
        if not engineer_ids:
            stats["projects_empty"] += 1
            continue
        stats["projects_with_engineers"] += 1
        object_title = proj.get("object_name") or proj.get("title") or proj["id"]
        contractor = proj.get("contractor") or ""

        for eid in engineer_ids:
            person = get_person(str(eid)) or {}
            fio = _norm_fio(person.get("fio") or str(eid))
            rows.append(
                {
                    "object": object_title,
                    "control_mode": _DEFAULT_CONTROL,
                    "contractor": contractor,
                    "position": person.get("position") or "",
                    "fio": fio,
                    "phone": _phone_digits(person.get("phone") or ""),
                }
            )

    return rows, stats


def _deployment_rows_from_month(month: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    from sk_reporter.personnel_store import get_person
    from sk_reporter.project_store import get_project

    rows: list[dict[str, Any]] = []
    stats = {"skipped_no_project": 0, "skipped_no_marks": 0}

    for person in month.get("people") or []:
        project_ids = person.get("project_ids") or []
        if not project_ids:
            stats["skipped_no_project"] += 1
            continue

        marks = person.get("marks") or []
        if marks and not any(m for m in marks):
            stats["skipped_no_marks"] += 1
            continue

        pinfo = get_person(person["person_id"]) if person.get("person_id") else None
        fio = _norm_fio(person.get("fio") or "")
        position = (pinfo or {}).get("position") or person.get("position") or ""
        phone = _phone_digits((pinfo or {}).get("phone") or "")

        for pid in project_ids:
            proj = get_project(pid) or {}
            rows.append(
                {
                    "object": proj.get("title") or pid,
                    "control_mode": _DEFAULT_CONTROL,
                    "contractor": proj.get("contractor") or "",
                    "position": position,
                    "fio": fio,
                    "phone": phone,
                }
            )

    return rows, stats


def _clear_data_rows(ws, start_row: int) -> None:
    for r in range(start_row, (ws.max_row or start_row) + 1):
        for c in range(1, 10):
            ws.cell(row=r, column=c).value = None


def _write_main_sheet(ws, rows: list[dict[str, Any]]) -> int:
    _clear_data_rows(ws, _DATA_START_ROW)
    for i, row in enumerate(rows, start=1):
        r = _DATA_START_ROW + i - 1
        ws.cell(row=r, column=1).value = i
        ws.cell(row=r, column=2).value = row["object"]
        ws.cell(row=r, column=3).value = row["control_mode"]
        ws.cell(row=r, column=4).value = row["contractor"] or None
        ws.cell(row=r, column=5).value = row["position"]
        ws.cell(row=r, column=6).value = row["fio"]
        ws.cell(row=r, column=7).value = row["phone"] or None
    return len(rows)


def _write_fio_sheet(ws, rows: list[dict[str, Any]]) -> None:
    counts = Counter(r["fio"] for r in rows if r.get("fio"))
    _clear_data_rows(ws, 2)
    for i, (fio, cnt) in enumerate(sorted(counts.items()), start=2):
        ws.cell(row=i, column=1).value = fio
        ws.cell(row=i, column=2).value = cnt


def _write_position_sheet(ws, rows: list[dict[str, Any]]) -> None:
    counts = Counter(r["position"] for r in rows if r.get("position"))
    _clear_data_rows(ws, 2)
    total = 0
    for i, (pos, cnt) in enumerate(sorted(counts.items(), key=lambda x: x[0]), start=2):
        ws.cell(row=i, column=1).value = pos
        ws.cell(row=i, column=2).value = cnt
        total += cnt
    if total:
        ws.cell(row=2 + len(counts), column=2).value = total


def _prepare_workbook_for_save(wb) -> None:
    for ws_sheet in wb.worksheets:
        ws_sheet.legacy_drawing = None


def _write_deployment_workbook(rows: list[dict[str, Any]], output: Path) -> int:
    template = deployment_template_path()
    shutil.copy2(template, output)

    wb = load_workbook(output, keep_vba=True)
    if _DEPLOY_SHEET not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Лист «{_DEPLOY_SHEET}» не найден в шаблоне расстановки")

    main_ws = wb[_DEPLOY_SHEET]
    written = _write_main_sheet(main_ws, rows)

    if _FIO_SHEET in wb.sheetnames:
        _write_fio_sheet(wb[_FIO_SHEET], rows)
    if _POS_SHEET in wb.sheetnames:
        _write_position_sheet(wb[_POS_SHEET], rows)

    _prepare_workbook_for_save(wb)
    wb.save(output)
    wb.close()
    return written


def projects_assignment_stats() -> dict[str, int]:
    from sk_reporter.project_store import list_projects_rich

    items = list_projects_rich()
    assignments = 0
    with_engineers = 0
    for p in items:
        ids = p.get("engineer_ids") or []
        if ids:
            with_engineers += 1
            assignments += len(ids)
    return {
        "projects_total": len(items),
        "projects_with_engineers": with_engineers,
        "projects_empty": len(items) - with_engineers,
        "assignments": assignments,
    }


def deployment_status() -> dict[str, Any]:
    manifest = load_luvr_manifest()
    template = None
    template_error = None
    data_rows = 0
    sheets: list[str] = []

    try:
        template = deployment_template_path()
    except FileNotFoundError as e:
        template_error = str(e)

    if template is not None:
        wb = load_workbook(template, read_only=True, data_only=True, keep_vba=True)
        sheets = list(wb.sheetnames)
        if _DEPLOY_SHEET in wb.sheetnames:
            ws = wb[_DEPLOY_SHEET]
            for r in range(_DATA_START_ROW, (ws.max_row or 0) + 1):
                if ws.cell(row=r, column=2).value or ws.cell(row=r, column=6).value:
                    data_rows += 1
        wb.close()

    return {
        "template_present": template is not None,
        "template_name": template.name if template else manifest.get("deployment"),
        "template_error": template_error,
        "sheet_name": _DEPLOY_SHEET,
        "sheets": sheets,
        "template_rows": data_rows,
        "output_dir": str(deployment_output_dir().relative_to(repo_root())),
        "assignments": projects_assignment_stats(),
    }


def build_deployment_from_projects() -> dict[str, Any]:
    rows, skip_stats = _deployment_rows_from_projects()
    if not rows:
        raise ValueError(
            "Нет назначений: отметьте инженеров на проектах в разделе «Проекты» и сохраните."
        )

    output = deployment_output_dir() / "Расстановка_справочник.xlsm"
    written = _write_deployment_workbook(rows, output)

    unique_fio = len({r["fio"] for r in rows})
    unique_projects = len({r["object"] for r in rows})

    return {
        "ok": True,
        "source": "projects",
        "output": str(output),
        "output_name": output.name,
        "download": f"/download/luvr/generated/{output.name}",
        "rows_written": written,
        "unique_people": unique_fio,
        "unique_objects": unique_projects,
        **skip_stats,
        "note": (
            "Из project.yaml: каждая пара инженер × объект. Подрядчик и водитель (кол. 8–9) — позже."
        ),
    }


def build_deployment_from_luvr(month_sheet: str) -> dict[str, Any]:
    from sk_reporter.luvr_store import luvr_month_payload

    month = luvr_month_payload(month_sheet)
    rows, skip_stats = _deployment_rows_from_month(month)
    if not rows:
        raise ValueError(
            f"Нет строк для расстановки по «{month_sheet}» "
            f"(без объекта: {skip_stats['skipped_no_project']}, без отметок: {skip_stats['skipped_no_marks']})"
        )

    year = month.get("year") or 2026
    safe_sheet = re.sub(r"[^\w\-]+", "_", month_sheet)
    output = deployment_output_dir() / f"Расстановка_{safe_sheet}_{year}.xlsm"
    written = _write_deployment_workbook(rows, output)

    unique_fio = len({r["fio"] for r in rows})
    unique_projects = len({r["object"] for r in rows})

    return {
        "ok": True,
        "source": "luvr",
        "sheet": month_sheet,
        "year": year,
        "output": str(output),
        "output_name": output.name,
        "download": f"/download/luvr/generated/{output.name}",
        "rows_written": written,
        "unique_people": unique_fio,
        "unique_objects": unique_projects,
        "skipped_no_project": skip_stats["skipped_no_project"],
        "skipped_no_marks": skip_stats["skipped_no_marks"],
        "note": (
            "Из ЛУВР за месяц (человек × объект, с отметками). "
            "Для черновика без ЛУВР — «Проекты» → «Сформировать расстановку»."
        ),
    }
