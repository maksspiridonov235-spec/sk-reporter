"""ЛУВР — лист учёта времени (xlsx ↔ luvr.yaml)."""

from __future__ import annotations

import re
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from sk_reporter.paths import luvr_dir

_MONTH_SHEETS = ("Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь")
_MONTH_NUM = {name: i + 1 for i, name in enumerate(_MONTH_SHEETS)}


def _personnel_fio_index() -> dict[str, str]:
    from sk_reporter.personnel_store import load_people

    return {_norm_fio(p["fio"]).lower(): p["id"] for p in load_people()}


def _manual_links_by_fio(data: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            if person.get("link_source") == "manual" and person.get("person_id"):
                out[_norm_fio(person["fio"]).lower()] = str(person["person_id"])
    return out


def _apply_person_link(person: dict[str, Any], by_fio: dict[str, str], manual_links: dict[str, str]) -> None:
    fio_key = _norm_fio(person["fio"]).lower()
    if fio_key in manual_links:
        person["person_id"] = manual_links[fio_key]
        person["link_source"] = "manual"
        return
    if person.get("link_source") == "manual" and person.get("person_id"):
        return
    pid = by_fio.get(fio_key)
    if pid:
        person["person_id"] = pid
        person["link_source"] = "auto"
    else:
        person.pop("person_id", None)
        person["link_source"] = "unmatched"


def enrich_luvr_links(data: dict[str, Any]) -> None:
    by_fio = _personnel_fio_index()
    manual_links = _manual_links_by_fio(data)
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            _apply_person_link(person, by_fio, manual_links)


def auto_link_luvr(save: bool = True) -> dict[str, Any]:
    data = load_luvr()
    if not data.get("months"):
        raise FileNotFoundError("luvr.yaml пуст")
    manual_links = _manual_links_by_fio(data)
    by_fio = _personnel_fio_index()
    linked = manual = unmatched = 0
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            _apply_person_link(person, by_fio, manual_links)
            src = person.get("link_source")
            if src == "manual":
                manual += 1
            elif src == "auto":
                linked += 1
            else:
                unmatched += 1
    if save:
        save_luvr(data)
    return {
        "ok": True,
        "linked": linked,
        "manual": manual,
        "unmatched": unmatched,
        "total": linked + manual + unmatched,
    }


def set_luvr_person_link(sheet: str, person_idx: int, person_id: str | None) -> dict[str, Any]:
    if sheet not in _MONTH_SHEETS:
        raise KeyError(f"Неизвестный лист: {sheet}")
    data = load_luvr()
    month = next((m for m in data.get("months") or [] if m.get("sheet") == sheet), None)
    if month is None:
        raise KeyError(f"Лист «{sheet}» не найден")
    people = month.get("people") or []
    if person_idx < 0 or person_idx >= len(people):
        raise IndexError("person_idx вне диапазона")

    person = people[person_idx]
    if person_id:
        from sk_reporter.personnel_store import get_person

        if get_person(person_id) is None:
            raise KeyError(f"person_id «{person_id}» не найден в справочнике сотрудников")
        person["person_id"] = person_id
        person["link_source"] = "manual"
    else:
        by_fio = _personnel_fio_index()
        fio_key = _norm_fio(person["fio"]).lower()
        pid = by_fio.get(fio_key)
        if pid:
            person["person_id"] = pid
            person["link_source"] = "auto"
        else:
            person.pop("person_id", None)
            person["link_source"] = "unmatched"

    if person.get("projects_source") != "manual":
        by_engineer = _engineer_projects_map()
        pid = person.get("person_id")
        if pid and pid in by_engineer:
            person["project_ids"] = list(by_engineer[pid])
            person["projects_source"] = "auto"
        else:
            person["project_ids"] = []
            person.pop("projects_source", None)

    save_luvr(data)
    return {
        "sheet": sheet,
        "person_idx": person_idx,
        "person_id": person.get("person_id"),
        "link_source": person.get("link_source"),
        "project_ids": person.get("project_ids") or [],
        "projects_source": person.get("projects_source"),
        "fio": person.get("fio"),
    }


def luvr_link_stats(months: list[dict[str, Any]]) -> dict[str, int]:
    linked = manual = unmatched = 0
    for month in months:
        for person in month.get("people") or []:
            src = person.get("link_source")
            if src == "manual":
                manual += 1
            elif person.get("person_id"):
                linked += 1
            else:
                unmatched += 1
    total = linked + manual + unmatched
    return {
        "total": total,
        "linked": linked + manual,
        "auto": linked,
        "manual": manual,
        "unmatched": unmatched,
    }


def luvr_personnel_options() -> list[dict[str, str]]:
    from sk_reporter.personnel_store import load_people

    return [{"id": p["id"], "fio": p["fio"]} for p in load_people()]


def luvr_projects_options() -> list[dict[str, str]]:
    from sk_reporter.project_store import list_projects_rich

    return [{"id": p["id"], "title": p.get("object_name") or p.get("title") or p["id"]} for p in list_projects_rich()]


def _engineer_projects_map() -> dict[str, list[str]]:
    from sk_reporter.project_store import engineer_project_map

    out: dict[str, list[str]] = {}
    for person_id, projs in engineer_project_map().items():
        out[str(person_id)] = [p["id"] for p in projs]
    return out


def _manual_projects_by_fio(data: dict[str, Any]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            if person.get("projects_source") == "manual" and person.get("project_ids"):
                out[_norm_fio(person["fio"]).lower()] = [str(x) for x in person["project_ids"]]
    return out


def _apply_person_projects(
    person: dict[str, Any],
    by_engineer: dict[str, list[str]],
    manual_projects: dict[str, list[str]],
) -> None:
    fio_key = _norm_fio(person["fio"]).lower()
    if fio_key in manual_projects:
        person["project_ids"] = list(manual_projects[fio_key])
        person["projects_source"] = "manual"
        return
    if person.get("projects_source") == "manual" and person.get("project_ids"):
        return
    person_id = person.get("person_id")
    if person_id and person_id in by_engineer:
        person["project_ids"] = list(by_engineer[person_id])
        person["projects_source"] = "auto"
    else:
        person["project_ids"] = []
        person.pop("projects_source", None)


def enrich_luvr_projects(data: dict[str, Any]) -> None:
    by_engineer = _engineer_projects_map()
    manual_projects = _manual_projects_by_fio(data)
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            _apply_person_projects(person, by_engineer, manual_projects)


def luvr_project_stats(months: list[dict[str, Any]]) -> dict[str, int]:
    with_projects = manual = empty = 0
    for month in months:
        for person in month.get("people") or []:
            ids = person.get("project_ids") or []
            if ids:
                with_projects += 1
                if person.get("projects_source") == "manual":
                    manual += 1
            else:
                empty += 1
    total = with_projects + empty
    return {
        "total": total,
        "with_projects": with_projects,
        "manual": manual,
        "auto": with_projects - manual,
        "empty": empty,
    }


def _validate_project_ids(project_ids: list[str]) -> list[str]:
    from sk_reporter.project_store import get_project

    cleaned: list[str] = []
    for raw in project_ids:
        pid = str(raw).strip()
        if not pid:
            continue
        if get_project(pid) is None:
            raise KeyError(f"project_id «{pid}» не найден")
        if pid not in cleaned:
            cleaned.append(pid)
    return cleaned


def set_luvr_person_projects(sheet: str, person_idx: int, project_ids: list[str]) -> dict[str, Any]:
    if sheet not in _MONTH_SHEETS:
        raise KeyError(f"Неизвестный лист: {sheet}")
    data = load_luvr()
    month = next((m for m in data.get("months") or [] if m.get("sheet") == sheet), None)
    if month is None:
        raise KeyError(f"Лист «{sheet}» не найден")
    people = month.get("people") or []
    if person_idx < 0 or person_idx >= len(people):
        raise IndexError("person_idx вне диапазона")

    person = people[person_idx]
    cleaned = _validate_project_ids(project_ids)
    if cleaned:
        person["project_ids"] = cleaned
        person["projects_source"] = "manual"
    else:
        person["project_ids"] = []
        person.pop("projects_source", None)
        by_engineer = _engineer_projects_map()
        pid = person.get("person_id")
        if pid and pid in by_engineer:
            person["project_ids"] = list(by_engineer[pid])
            person["projects_source"] = "auto"

    save_luvr(data)
    return {
        "sheet": sheet,
        "person_idx": person_idx,
        "project_ids": person.get("project_ids") or [],
        "projects_source": person.get("projects_source"),
        "fio": person.get("fio"),
    }


def auto_assign_luvr_projects(save: bool = True) -> dict[str, Any]:
    data = load_luvr()
    if not data.get("months"):
        raise FileNotFoundError("luvr.yaml пуст")
    manual_projects = _manual_projects_by_fio(data)
    by_engineer = _engineer_projects_map()
    auto = manual = empty = 0
    for month in data.get("months") or []:
        for person in month.get("people") or []:
            _apply_person_projects(person, by_engineer, manual_projects)
            src = person.get("projects_source")
            if src == "manual":
                manual += 1
            elif person.get("project_ids"):
                auto += 1
            else:
                empty += 1
    if save:
        save_luvr(data)
    return {
        "ok": True,
        "auto": auto,
        "manual": manual,
        "empty": empty,
        "total": auto + manual + empty,
    }


def _luvr_xlsx() -> Path:
    folder = luvr_dir()
    for name in ("ЛУВР.xlsx", "luvr.xlsx"):
        p = folder / name
        if p.is_file():
            return p
    matches = sorted(folder.glob("*.xlsx"))
    if matches:
        return matches[0]
    raise FileNotFoundError(f"Нет xlsx в {folder}")


def _cell_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _norm_fio(v: Any) -> str:
    return " ".join(str(v or "").split())


def _count_days(values: list[Any]) -> dict[str, int]:
    out = {"present": 0, "half": 0, "other": 0}
    for v in values:
        if v is None or v == "":
            continue
        s = str(v).strip()
        if s == "1":
            out["present"] += 1
        elif s in {"0.5", "0,5"}:
            out["half"] += 1
        else:
            out["other"] += 1
    return out


def _norm_mark(v: Any) -> str:
    if v is None or v == "":
        return ""
    s = str(v).strip().replace(",", ".")
    if s == "1":
        return "1"
    if s == "0.5":
        return "0.5"
    return s


def _mark_to_cell_value(mark: str) -> Any:
    if not mark:
        return None
    if mark == "1":
        return 1
    if mark == "0.5":
        return 0.5
    return mark


def _cell_values_equal(current: Any, new: Any) -> bool:
    if current is None and new is None:
        return True
    if current is None or new is None:
        return False
    if isinstance(current, float) and isinstance(new, float):
        return abs(current - new) < 1e-9
    return _norm_mark(current) == _norm_mark(new)


def _is_mark_like(v: Any) -> bool:
    if v is None or v == "":
        return True
    s = _norm_mark(v)
    return s in {"", "1", "0.5"}


def _infer_mark_columns_ws(ws, data_row_start: int, limit: int | None = None) -> list[int]:
    cols: list[int] = []
    max_col = ws.max_column or 0
    for c in range(6, max_col + 1):
        v = ws.cell(row=data_row_start, column=c).value
        if not _is_mark_like(v):
            if cols:
                break
            continue
        cols.append(c)
        if limit is not None and len(cols) >= limit:
            break
        if limit is None and len(cols) >= 31:
            break
    return cols


def _day_cols_for_month(ws, layout: dict[str, Any], month: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    yaml_days = (month or {}).get("days") or []
    if yaml_days and all(d.get("col") for d in yaml_days):
        return [{"col": int(d["col"]), "date": d.get("date"), "day": d.get("day")} for d in yaml_days]

    day_row = layout["hdr_row"] + 1
    day_cols: list[dict[str, Any]] = []
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        d = _cell_date(ws.cell(row=day_row, column=col).value)
        if d:
            day_cols.append({"col": col, "date": d.isoformat(), "day": d.day})

    if yaml_days and len(day_cols) == len(yaml_days):
        return day_cols

    if yaml_days:
        return [
            {
                "col": 6 + i,
                "date": yd.get("date") or (day_cols[i]["date"] if i < len(day_cols) else None),
                "day": yd.get("day") if yd.get("day") is not None else (i + 1),
            }
            for i, yd in enumerate(yaml_days)
        ]

    mark_cols = _infer_mark_columns_ws(ws, layout["data_row_start"])
    if mark_cols:
        return [{"col": c, "date": None, "day": i + 1} for i, c in enumerate(mark_cols)]
    return day_cols


def _layout_for_sheet_ws(ws) -> dict[str, Any] | None:
    hdr_row = None
    max_scan = min(ws.max_row or 0, 80)
    for r in range(1, max_scan + 1):
        if ws.cell(row=r, column=2).value == "ФИО":
            hdr_row = r
            break
    if hdr_row is None:
        return None

    day_row = hdr_row + 1
    day_cols: list[dict[str, Any]] = []
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        d = _cell_date(ws.cell(row=day_row, column=col).value)
        if d:
            day_cols.append({"col": col, "date": d.isoformat(), "day": d.day})

    if not day_cols:
        mark_cols = _infer_mark_columns_ws(ws, hdr_row + 2)
        for i, col in enumerate(mark_cols):
            day_cols.append({"col": col, "date": None, "day": i + 1})

    return {
        "hdr_row": hdr_row,
        "data_row_start": hdr_row + 2,
        "day_cols": day_cols,
    }


def _person_row_by_fio_ws(ws, data_row_start: int, fio: str) -> int | None:
    target = _norm_fio(fio)
    for r in range(data_row_start, (ws.max_row or 0) + 1):
        if _norm_fio(ws.cell(row=r, column=2).value) == target:
            return r
    return None


def _parse_sheet(ws) -> dict[str, Any] | None:
    layout = _layout_for_sheet_ws(ws)
    if layout is None:
        return None

    hdr_row = layout["hdr_row"]
    day_cols = layout["day_cols"]
    data_row_start = layout["data_row_start"]

    title = ""
    for r in range(1, hdr_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and "Лист учета" in val:
            title = val.replace("\n", " ").strip()
            break

    people: list[dict[str, Any]] = []
    for r in range(data_row_start, (ws.max_row or 0) + 1):
        fio = _norm_fio(ws.cell(row=r, column=2).value)
        if not fio:
            continue
        day_values = [ws.cell(row=r, column=c["col"]).value for c in day_cols]
        counts = _count_days(day_values)
        marks = [_norm_mark(v) for v in day_values]
        people.append(
            {
                "num": ws.cell(row=r, column=1).value,
                "fio": fio,
                "position": str(ws.cell(row=r, column=3).value or "").strip(),
                "nrs": str(ws.cell(row=r, column=4).value or "").strip(),
                "specialty": str(ws.cell(row=r, column=5).value or "").strip(),
                "days_present": counts["present"],
                "days_half": counts["half"],
                "days_marked": counts["present"] + counts["half"] + counts["other"],
                "marks": marks,
            }
        )

    month_num = _MONTH_NUM.get(ws.title)
    year = day_cols[0]["date"][:4] if day_cols else None
    return {
        "sheet": ws.title,
        "year": int(year) if year else None,
        "month": month_num,
        "title": title,
        "people_count": len(people),
        "days_in_sheet": len(day_cols),
        "days": [{"date": c["date"], "day": c["day"], "col": c["col"]} for c in day_cols],
        "people": people,
    }


def export_luvr() -> Path:
    xlsx = _luvr_xlsx()
    wb = load_workbook(xlsx, data_only=True)
    months: list[dict[str, Any]] = []
    contract = ""
    for sheet_name in wb.sheetnames:
        if sheet_name not in _MONTH_SHEETS:
            continue
        parsed = _parse_sheet(wb[sheet_name])
        if parsed and parsed["people_count"]:
            if not contract and parsed.get("title"):
                m = re.search(r"договор[^\d]*([\w.\-]+)", parsed["title"], re.I)
                if m:
                    contract = m.group(1)
            months.append(parsed)
    wb.close()

    old = load_luvr()
    manual_links = _manual_links_by_fio(old)
    manual_projects = _manual_projects_by_fio(old)

    payload = {
        "source": xlsx.name,
        "source_kb": round(xlsx.stat().st_size / 1024, 1),
        "contract": contract,
        "xlsx_stale": False,
        "months": months,
    }
    by_fio = _personnel_fio_index()
    by_engineer = _engineer_projects_map()
    for month in payload["months"]:
        for person in month.get("people") or []:
            _apply_person_link(person, by_fio, manual_links)
            _apply_person_projects(person, by_engineer, manual_projects)

    out = luvr_dir() / "luvr.yaml"
    out.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
    load_luvr.cache_clear()
    return out


def import_luvr_from_xlsx() -> dict[str, Any]:
    out = export_luvr()
    data = load_luvr()
    return {
        "ok": True,
        "yaml": str(out),
        "source": data.get("source"),
        "months_count": len(data.get("months") or []),
        "xlsx_stale": False,
    }


def _sync_month_to_ws(ws, month: dict[str, Any], layout: dict[str, Any]) -> int:
    updated = 0
    day_cols = _day_cols_for_month(ws, layout, month)
    if not day_cols:
        return 0
    for person in month.get("people") or []:
        row = _person_row_by_fio_ws(ws, layout["data_row_start"], person["fio"])
        if row is None:
            continue
        marks = person.get("marks") or []
        for i, dc in enumerate(day_cols):
            if i >= len(marks):
                break
            new_val = _mark_to_cell_value(marks[i])
            cell = ws.cell(row=row, column=dc["col"])
            if not _cell_values_equal(cell.value, new_val):
                cell.value = new_val
                updated += 1
    return updated


def sync_luvr_to_xlsx(sheet: str | None = None) -> dict[str, Any]:
    xlsx = _luvr_xlsx()
    data = load_luvr()
    months = data.get("months") or []
    if not months:
        raise FileNotFoundError("luvr.yaml пуст")

    wb = load_workbook(xlsx)
    total_updated = 0
    sheets_synced: list[str] = []

    for month in months:
        name = month.get("sheet")
        if not name or name not in wb.sheetnames:
            continue
        if sheet is not None and name != sheet:
            continue
        layout = _layout_for_sheet_ws(wb[name])
        if not layout:
            continue
        n = _sync_month_to_ws(wb[name], month, layout)
        total_updated += n
        sheets_synced.append(name)

    if sheet is not None and sheet not in sheets_synced:
        wb.close()
        raise KeyError(f"Лист «{sheet}» не найден или пуст")

    wb.save(xlsx)
    wb.close()

    data["xlsx_stale"] = False
    data["source"] = xlsx.name
    data["source_kb"] = round(xlsx.stat().st_size / 1024, 1)
    save_luvr(data)

    return {
        "ok": True,
        "xlsx": xlsx.name,
        "cells_updated": total_updated,
        "sheets": sheets_synced,
        "xlsx_stale": False,
    }


def write_luvr_mark_to_xlsx(sheet: str, person_idx: int, day_idx: int, mark: str) -> bool:
    return write_luvr_marks_batch_to_xlsx(sheet, [(person_idx, day_idx, mark)])


def write_luvr_marks_batch_to_xlsx(sheet: str, updates: list[tuple[int, int, str]]) -> bool:
    if not updates:
        return True
    try:
        xlsx = _luvr_xlsx()
    except FileNotFoundError:
        return False

    data = load_luvr()
    month = next((m for m in data.get("months") or [] if m.get("sheet") == sheet), None)
    if month is None:
        return False
    people = month.get("people") or []
    days = month.get("days") or []

    wb = load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[sheet]
    layout = _layout_for_sheet_ws(ws)
    if not layout:
        wb.close()
        return False

    day_cols = _day_cols_for_month(ws, layout, month)
    changed = False
    for person_idx, day_idx, mark in updates:
        if person_idx < 0 or person_idx >= len(people):
            continue
        if day_idx < 0 or day_idx >= len(day_cols):
            continue
        row = _person_row_by_fio_ws(ws, layout["data_row_start"], people[person_idx]["fio"])
        if row is None:
            continue
        col = day_cols[day_idx]["col"]
        new_val = _mark_to_cell_value(_norm_mark(mark))
        cell = ws.cell(row=row, column=col)
        if not _cell_values_equal(cell.value, new_val):
            cell.value = new_val
            changed = True

    if changed:
        wb.save(xlsx)
    wb.close()
    return True


def _recalc_person_stats(person: dict[str, Any]) -> None:
    marks = person.get("marks") or []
    counts = _count_days(marks)
    person["days_present"] = counts["present"]
    person["days_half"] = counts["half"]
    person["days_marked"] = counts["present"] + counts["half"] + counts["other"]


def save_luvr(data: dict[str, Any] | None = None) -> Path:
    payload = data if data is not None else load_luvr()
    out = luvr_dir() / "luvr.yaml"
    out.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
    load_luvr.cache_clear()
    return out


def update_luvr_mark(sheet: str, person_idx: int, day_idx: int, mark: Any) -> dict[str, Any]:
    if sheet not in _MONTH_SHEETS:
        raise KeyError(f"Неизвестный лист: {sheet}")
    data = load_luvr()
    if not data.get("months"):
        raise FileNotFoundError("luvr.yaml пуст — сначала build_engineer_data.py --luvr")

    month = next((m for m in data["months"] if m.get("sheet") == sheet), None)
    if month is None:
        raise KeyError(f"Лист «{sheet}» не найден в luvr.yaml")

    people = month.get("people") or []
    if person_idx < 0 or person_idx >= len(people):
        raise IndexError("person_idx вне диапазона")

    days = month.get("days") or []
    if day_idx < 0 or day_idx >= len(days):
        raise IndexError("day_idx вне диапазона")

    person = people[person_idx]
    marks = list(person.get("marks") or [])
    while len(marks) < len(days):
        marks.append("")

    norm = _norm_mark(mark)
    marks[day_idx] = norm
    person["marks"] = marks
    _recalc_person_stats(person)

    xlsx_synced = write_luvr_mark_to_xlsx(sheet, person_idx, day_idx, norm)
    data["xlsx_stale"] = not xlsx_synced
    save_luvr(data)

    return {
        "sheet": sheet,
        "person_idx": person_idx,
        "day_idx": day_idx,
        "mark": norm,
        "days_present": person["days_present"],
        "days_marked": person["days_marked"],
        "xlsx_synced": xlsx_synced,
        "xlsx_stale": data.get("xlsx_stale", False),
    }


def update_luvr_marks_batch(sheet: str, updates: list[dict[str, Any]]) -> dict[str, Any]:
    if sheet not in _MONTH_SHEETS:
        raise KeyError(f"Неизвестный лист: {sheet}")
    if not updates:
        return {"sheet": sheet, "updated": 0, "people": [], "xlsx_synced": True, "xlsx_stale": False}
    if len(updates) > 5000:
        raise ValueError("Слишком много ячеек за один запрос (макс. 5000)")

    data = load_luvr()
    if not data.get("months"):
        raise FileNotFoundError("luvr.yaml пуст — сначала build_engineer_data.py --luvr")

    month = next((m for m in data["months"] if m.get("sheet") == sheet), None)
    if month is None:
        raise KeyError(f"Лист «{sheet}» не найден в luvr.yaml")

    people = month.get("people") or []
    days = month.get("days") or []
    day_count = len(days)

    merged: dict[tuple[int, int], str] = {}
    for raw in updates:
        person_idx = int(raw["person_idx"])
        day_idx = int(raw["day_idx"])
        if person_idx < 0 or person_idx >= len(people):
            raise IndexError(f"person_idx {person_idx} вне диапазона")
        if day_idx < 0 or day_idx >= day_count:
            raise IndexError(f"day_idx {day_idx} вне диапазона")
        merged[(person_idx, day_idx)] = _norm_mark(raw.get("mark"))

    xlsx_updates: list[tuple[int, int, str]] = []
    touched_people: set[int] = set()

    for (person_idx, day_idx), norm in merged.items():
        person = people[person_idx]
        marks = list(person.get("marks") or [])
        while len(marks) < day_count:
            marks.append("")
        if marks[day_idx] == norm:
            continue
        marks[day_idx] = norm
        person["marks"] = marks
        touched_people.add(person_idx)
        xlsx_updates.append((person_idx, day_idx, norm))

    for person_idx in touched_people:
        _recalc_person_stats(people[person_idx])

    xlsx_synced = write_luvr_marks_batch_to_xlsx(sheet, xlsx_updates) if xlsx_updates else True
    data["xlsx_stale"] = not xlsx_synced if xlsx_updates else data.get("xlsx_stale", False)
    save_luvr(data)

    people_stats = [
        {
            "person_idx": pi,
            "days_present": people[pi]["days_present"],
            "days_marked": people[pi]["days_marked"],
        }
        for pi in sorted(touched_people)
    ]

    return {
        "sheet": sheet,
        "updated": len(xlsx_updates),
        "people": people_stats,
        "xlsx_synced": xlsx_synced,
        "xlsx_stale": data.get("xlsx_stale", False),
    }


@lru_cache(maxsize=1)
def load_luvr() -> dict[str, Any]:
    cache = luvr_dir() / "luvr.yaml"
    if cache.is_file():
        return yaml.safe_load(cache.read_text(encoding="utf-8")) or {}
    return {}


def _cache_has_grid(months: list[dict[str, Any]]) -> bool:
    if not months:
        return False
    m = months[0]
    if not m.get("days"):
        return False
    people = m.get("people") or []
    return bool(people and "marks" in people[0])


def luvr_planning_payload() -> dict[str, Any]:
    from sk_reporter.paths import repo_root

    folder = luvr_dir()
    xlsx_path = None
    try:
        xlsx_path = _luvr_xlsx()
    except FileNotFoundError:
        pass

    data = load_luvr()
    months = data.get("months") or []
    if xlsx_path is not None and (not months or not _cache_has_grid(months)):
        try:
            export_luvr()
            data = load_luvr()
            months = data.get("months") or []
        except Exception:
            pass

    enrich_luvr_links(data)
    enrich_luvr_projects(data)
    months = data.get("months") or []

    yaml_path = luvr_dir() / "luvr.yaml"
    default_month = months[-1]["sheet"] if months else None
    xlsx_stale = bool(data.get("xlsx_stale")) if xlsx_path else False
    link_stats = luvr_link_stats(months)
    project_stats = luvr_project_stats(months)

    appendix7: dict[str, Any] = {}
    try:
        from sk_reporter.appendix7_store import appendix7_status

        appendix7 = appendix7_status()
    except Exception:
        appendix7 = {"template_present": False}

    deployment: dict[str, Any] = {}
    try:
        from sk_reporter.deployment_store import deployment_status

        deployment = deployment_status()
    except Exception:
        deployment = {"template_present": False}

    return {
        "folder": str(folder.relative_to(repo_root())),
        "source": data.get("source") or (xlsx_path.name if xlsx_path else None),
        "source_kb": data.get("source_kb") or (round(xlsx_path.stat().st_size / 1024, 1) if xlsx_path else None),
        "contract": data.get("contract"),
        "months": months,
        "default_month": default_month,
        "cache_ready": bool(months),
        "cache_from_yaml": bool(months) and yaml_path.is_file(),
        "grid_ready": _cache_has_grid(months),
        "editable": _cache_has_grid(months),
        "xlsx_present": xlsx_path is not None,
        "xlsx_stale": xlsx_stale,
        "personnel": luvr_personnel_options(),
        "projects": luvr_projects_options(),
        "link_stats": link_stats,
        "project_stats": project_stats,
        "appendix7": appendix7,
        "deployment": deployment,
    }


def luvr_month_payload(sheet: str) -> dict[str, Any]:
    data = load_luvr()
    for m in data.get("months") or []:
        if m.get("sheet") == sheet:
            return m
    raise KeyError(sheet)
