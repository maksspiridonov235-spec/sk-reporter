"""
generate_rasstanovka.py
Формирует Отчёт-расстановка ДД.ММ.ГГГГ.html из Приложения 7.

Источники:
  - Приложение 7 (xlsm) → ФИО инженера, Объект, Подрядчик
  - справочник.json      → Должность, Телефон, Режим контроля
  - Рабочая исходник     → ФИО и тел. закреплённого
"""

import os
import traceback
from html import escape
from datetime import datetime

import openpyxl

_DEFAULT_REZHIM = 'Инспекционный контроль, Проверка ИТД'
SHEET_RAB = 'Рабочая исходник'

_DOLZH_ORDER = [
    'Инженер СК (общестроительные работы, сварочные технологии)',
    'Инженер СК (общестроительные работы)',
    'Инженер СК (электромонтажные работы, КИПиА)',
    'Инженер ПИЛ  (НК, УЗК)',
    'Инженер СЛ НК',
    'Инженер ОЗОТОБОС СПД',
    'Инженер ОЗОТОБОС',
    'Инженер БДД',
    'Инженер СК (Геодезические работы)',
    'Инженер СК (Геологоизыскательские работы)',
    'Системотехник / Инженер ПТО',
    'Инженер ПТО СПД СГМ',
    'Инженер ПТО СГМ',
    'Инженер ПТО Тюмень',
    'Руководитель СК',
]
_DOLZH_IDX = {d: i for i, d in enumerate(_DOLZH_ORDER)}


def _dolzh_sort_key(dolzh: str) -> int:
    try:
        from sk_reporter.position_db import position_sort_map

        sm = position_sort_map()
        if sm and dolzh in sm:
            return sm[dolzh]
    except Exception:
        pass
    return _DOLZH_IDX.get(dolzh, 999)


# ─────────────────────────────────────────────────────────────────────────────
from sk_reporter.deployment.lookup import load_sprav_dict as _load_sprav


def _read_rabochaya(template_path):
    """Читает Рабочая исходник → {ФИО: {zakr_fio, zakr_tel}}."""
    if not template_path or not os.path.exists(template_path):
        return {}
    try:
        wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=True)
        if SHEET_RAB not in wb.sheetnames:
            return {}
        ws = wb[SHEET_RAB]
        result = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            fio = row[2]
            if not fio or not isinstance(fio, str):
                continue
            fio = fio.strip()
            if not fio or fio == 'ФИО':
                continue
            result[fio] = {
                'zakr_fio': str(row[6]).strip() if row[6] else '',
                'zakr_tel': str(row[7]).strip() if row[7] else '',
            }
        return result
    except Exception:
        return {}


def _parse_excel_date(v):
    """Пробует преобразовать значение ячейки Excel в date. None если не дата."""
    if v is None:
        return None
    try:
        if hasattr(v, 'date'):
            return v.date()
        if isinstance(v, (int, float)) and v > 20000:
            return (datetime(1899, 12, 30) + __import__('datetime').timedelta(days=int(v))).date()
    except Exception:
        pass
    return None


def _is_date_like(v):
    """Грубая проверка — похоже ли значение на дату/число дня."""
    if v is None:
        return False
    if hasattr(v, 'date'):
        return True
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv > 20000 or (1 <= iv <= 31)
    if isinstance(v, str):
        s = v.strip().replace('.', '').replace(',', '')
        if s.isdigit():
            return 1 <= int(s) <= 31
    return False


def _read_pril7_openpyxl(pril7_path, log_func, override_date=None):
    """Читает Приложение 7 через openpyxl — без COM, безопасно в потоках Flask."""
    rows = []
    try:
        wb = openpyxl.load_workbook(pril7_path, keep_vba=True, data_only=True)
        if 'Отчет о ВОУ' not in wb.sheetnames:
            log_func('  ОШИБКА: лист "Отчет о ВОУ" не найден')
            return rows
        ws = wb['Отчет о ВОУ']

        # 1. Дата отчёта
        if override_date:
            report_date = override_date
            log_func(f'  Дата (из интерфейса): {report_date.strftime("%d.%m.%Y")}')
        else:
            report_date = _parse_excel_date(ws.cell(4, 48).value)
            if report_date:
                log_func(f'  Дата (из AV4): {report_date.strftime("%d.%m.%Y")}')
            else:
                log_func('  ПРЕДУПРЕЖДЕНИЕ: дата в AV4 не найдена — берём все строки с ФИО и Объектом')

        # 2. Строка заголовков с датами (строки 7–16)
        header_row, max_count = 0, 0
        for r in range(7, 17):
            cnt = sum(1 for c in range(6, 61) if _is_date_like(ws.cell(r, c).value))
            if cnt > max_count:
                max_count, header_row = cnt, r
        if header_row == 0:
            header_row = 12
        log_func(f'  Строка дат: {header_row}')

        # 3. Столбец нужной даты
        today_col = None
        if report_date:
            for c in range(6, 61):
                d = _parse_excel_date(ws.cell(header_row, c).value)
                if d and d == report_date:
                    today_col = c
                    break
                v = ws.cell(header_row, c).value
                if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
                    if int(v) == report_date.day:
                        today_col = c
                        break
            if today_col:
                log_func(f'  Столбец даты: {today_col}')
            else:
                log_func('  Столбец даты не найден — берём все строки')

        # 4. Сбор данных
        data_start = header_row + 1
        max_row = ws.max_row
        log_func(f'  Читаю строки {data_start}–{max_row}')

        for r in range(data_start, max_row + 1):
            fio_val = ws.cell(r, 5).value   # E
            obj_val = ws.cell(r, 3).value   # C
            if not fio_val or not obj_val:
                continue
            if today_col and not ws.cell(r, today_col).value:
                continue
            rows.append({
                'fio':        str(fio_val).strip(),
                'obj':        str(obj_val).strip(),
                'contractor': str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else '',
            })

        log_func(f'  Найдено строк: {len(rows)}')
        wb.close()

    except Exception as e:
        log_func(f'  ОШИБКА чтения Приложения 7: {e}')
        import traceback as tb
        log_func(tb.format_exc())
    return rows


def _build_html(rows_data, today, groups):

    # ── Основная таблица расстановки ─────────────────────────────────────────
    main_rows = []
    num = 0
    for _, group in groups:
        for d in group:
            num += 1
            bg = ' style="background:#f0f6ff"' if num % 2 else ''
            zakr = ' '.join(filter(None, [d['zakr_fio'], d['zakr_tel']]))
            main_rows.append(
                f'<tr{bg}>'
                f'<td style="text-align:center">{num}</td>'
                f'<td>{escape(d["obj"])}</td>'
                f'<td>{escape(d["rezhim"])}</td>'
                f'<td>{escape(d["contractor"])}</td>'
                f'<td>{escape(d["dolzh"])}</td>'
                f'<td style="white-space:nowrap">{escape(d["fio"])}</td>'
                f'<td style="white-space:nowrap">{escape(d["tel"])}</td>'
                f'<td>{escape(zakr)}</td>'
                f'</tr>'
            )

    # ── Итоговые строки в конце основной таблицы ─────────────────────────────
    total_eng_main = len({d['fio'] for d in rows_data})
    total_obj_main = len({d['obj'] for d in rows_data if d['obj']})
    main_rows.append(
        f'<tr style="background:#bdd7ee;font-weight:bold;">'
        f'<td colspan="5" style="text-align:right">Итого персонала:</td>'
        f'<td style="text-align:center" colspan="3">{total_eng_main}</td>'
        f'</tr>'
    )
    main_rows.append(
        f'<tr style="background:#bdd7ee;font-weight:bold;">'
        f'<td colspan="5" style="text-align:right">Итого объектов:</td>'
        f'<td style="text-align:center" colspan="3">{total_obj_main}</td>'
        f'</tr>'
    )

    # ── Итоги по должностям (для вкладки Итоги) ──────────────────────────────
    total_eng = 0
    itog_rows = []
    for i, (dolzh, group) in enumerate(groups):
        grp_eng = len({d['fio'] for d in group})
        grp_obj = len({d['obj'] for d in group if d['obj']})
        total_eng += grp_eng
        bg = ' style="background:#f0f6ff"' if i % 2 else ''
        itog_rows.append(
            f'<tr{bg}>'
            f'<td style="text-align:center">{i+1}</td>'
            f'<td>{escape(dolzh or "—")}</td>'
            f'<td style="text-align:center">{grp_eng}</td>'
            f'<td style="text-align:center">{grp_obj}</td>'
            f'</tr>'
        )
    total_unique_obj = len({d['obj'] for d in rows_data if d['obj']})
    itog_rows.append(
        f'<tr style="background:#bdd7ee;font-weight:bold;">'
        f'<td colspan="2">Итого персонала:</td>'
        f'<td style="text-align:center">{total_eng}</td>'
        f'<td></td></tr>'
    )
    itog_rows.append(
        f'<tr style="background:#bdd7ee;font-weight:bold;">'
        f'<td colspan="2">Итого объектов:</td>'
        f'<td></td>'
        f'<td style="text-align:center">{total_unique_obj}</td>'
        f'</tr>'
    )

    main_body = '\n'.join(main_rows)
    itog_body = '\n'.join(itog_rows)

    return f'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Отчёт-расстановка {today}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h2 {{ font-size: 13px; text-align: center; margin-bottom: 10px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #999; padding: 4px 6px; vertical-align: top; }}
  th {{ background: #c6d9f1; text-align: center; font-size: 11px; }}
  .tabs {{ display:flex; gap:4px; margin-bottom:12px; border-bottom:2px solid #c6d9f1; }}
  .tab-btn {{ padding:6px 18px; border:none; background:none; cursor:pointer;
              font-size:12px; color:#555; border-bottom:2px solid transparent;
              margin-bottom:-2px; transition:color .2s; }}
  .tab-btn.active {{ color:#1a5db5; border-bottom-color:#1a5db5; font-weight:600; }}
  .tab-pane {{ display:none; }}
  .tab-pane.active {{ display:block; }}
  @media print {{ .tabs {{ display:none; }} .tab-pane {{ display:block !important; }} }}
</style>
</head>
<body>
<h2>Отчёт-расстановка на {today}</h2>

<div class="tabs">
  <button class="tab-btn active" onclick="switchTab('main',this)">Расстановка</button>
  <button class="tab-btn" onclick="switchTab('itog',this)">Итоги</button>
</div>

<div id="tab-main" class="tab-pane active">
<table>
<thead><tr>
  <th>№ п/п</th>
  <th>Наименование объектов</th>
  <th>Режим контроля</th>
  <th>Наименование подрядной организации</th>
  <th>Должность СК</th>
  <th>ФИО закреплённого специалиста СК</th>
  <th>Телефон</th>
  <th>Закреплённое за специалистом строительного контроля, ФИО водителя, телефон</th>
</tr></thead>
<tbody>{main_body}</tbody>
</table>
</div>

<div id="tab-itog" class="tab-pane">
<table style="width:auto;min-width:420px;">
<thead><tr>
  <th>№</th><th>Должность</th><th>Инженеров</th><th>Объектов</th>
</tr></thead>
<tbody>{itog_body}</tbody>
</table>
</div>

<script>
function switchTab(name, btn) {{
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  btn.classList.add('active');
}}
</script>
</body>
</html>'''


def _prepare_data(pril7_path, template_path, report_date, log_func):
    """Общая подготовка данных. Возвращает (rows_data, today_str) или (None, None)."""
    override_date = None
    if report_date:
        try:
            override_date = datetime.strptime(report_date, '%Y-%m-%d').date()
            log_func(f'Дата из интерфейса: {override_date.strftime("%d.%m.%Y")}')
        except ValueError:
            log_func(f'Неверный формат даты: {report_date}, читаю из AV4')

    log_func('Читаю справочник...')
    sprav = _load_sprav()

    log_func('Читаю Рабочая исходник из шаблона...')
    rab_map = _read_rabochaya(template_path)

    log_func('Читаю Приложение 7...')
    pril7_rows = _read_pril7_openpyxl(pril7_path, log_func, override_date=override_date)

    if not pril7_rows:
        log_func('ОШИБКА: нет данных в Приложении 7')
        return None, None

    log_func(f'Найдено строк: {len(pril7_rows)}')
    today = override_date.strftime('%d.%m.%Y') if override_date else datetime.now().strftime('%d.%m.%Y')

    rows_data = []
    for item in pril7_rows:
        fio  = item['fio']
        sp   = sprav.get(fio, {})
        rb   = rab_map.get(fio, {})
        rows_data.append({
            'obj':        item['obj'],
            'rezhim':     sp.get('rezhim', '').strip() or _DEFAULT_REZHIM,
            'contractor': item['contractor'],
            'dolzh':      sp.get('dolzhnost', ''),
            'fio':        fio,
            'tel':        sp.get('telefon', ''),
            'zakr_fio':   rb.get('zakr_fio', ''),
            'zakr_tel':   rb.get('zakr_tel', ''),
        })
    return rows_data, today


def _build_excel(rows_data, today, groups):
    """Создаёт openpyxl Workbook с листами Расстановка и Итоги."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_font  = Font(bold=True)
    hdr_fill  = PatternFill('solid', fgColor='C6D9F1')
    grp_fill  = PatternFill('solid', fgColor='DCE6F1')
    tot_fill  = PatternFill('solid', fgColor='BDD7EE')
    alt_fill  = PatternFill('solid', fgColor='F0F6FF')
    thin      = Side(style='thin', color='999999')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='top', wrap_text=True)
    wrap      = Alignment(vertical='top', wrap_text=True)

    def _style(cell, font=None, fill=None, align=None):
        cell.border = border
        if font:  cell.font  = font
        if fill:  cell.fill  = fill
        if align: cell.alignment = align

    # ── Лист 1: Расстановка ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Расстановка'

    # Строка 1 — заголовок документа
    ws1.append([f'Отчёт-расстановка на {today}'])
    ws1.merge_cells('A1:H1')
    title_cell = ws1.cell(1, 1)
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 22

    # Строка 2 — шапка таблицы
    headers = [
        '№ п/п',
        'Наименование объектов',
        'Режим контроля',
        'Наименование подрядной организации',
        'Должность СК',
        'ФИО закреплённого специалиста СК',
        'Телефон',
        'Закреплённое за специалистом строительного контроля, ФИО водителя, телефон',
    ]
    ws1.append(headers)
    for cell in ws1[2]:
        _style(cell, font=hdr_font, fill=hdr_fill, align=center)

    num = 0
    for _, group in groups:
        for d in group:
            num += 1
            zakr = ' '.join(filter(None, [d['zakr_fio'], d['zakr_tel']]))
            ws1.append([num, d['obj'], d['rezhim'], d['contractor'],
                        d['dolzh'], d['fio'], d['tel'], zakr])
            row = ws1.max_row
            fill = alt_fill if num % 2 else None
            for cell in ws1[row]:
                _style(cell, fill=fill, align=wrap)
            ws1.cell(row, 1).alignment = center

    # Строки итогов
    total_eng_main = len({d['fio'] for d in rows_data})
    total_obj_main = len({d['obj'] for d in rows_data if d['obj']})
    for label, val in [('Итого персонала:', total_eng_main),
                        ('Итого объектов:',  total_obj_main)]:
        ws1.append(['', '', '', '', label, val, '', ''])
        row = ws1.max_row
        ws1.merge_cells(f'A{row}:D{row}')
        ws1.merge_cells(f'F{row}:H{row}')
        for cell in ws1[row]:
            _style(cell, font=hdr_font, fill=tot_fill)
        ws1.cell(row, 5).alignment = Alignment(horizontal='right',  vertical='center')
        ws1.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')

    col_widths = [5, 35, 28, 35, 35, 30, 15, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A3'

    # ── Лист 2: Итоги ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Итоги')
    ws2.append(['№', 'Должность', 'Инженеров', 'Объектов'])
    for cell in ws2[1]:
        _style(cell, font=hdr_font, fill=hdr_fill, align=center)

    total_eng = 0
    for i, (dolzh, group) in enumerate(groups):
        grp_eng = len({d['fio'] for d in group})
        grp_obj = len({d['obj'] for d in group if d['obj']})
        total_eng += grp_eng
        ws2.append([i + 1, dolzh or '—', grp_eng, grp_obj])
        row = ws2.max_row
        fill = alt_fill if i % 2 else None
        for cell in ws2[row]:
            _style(cell, fill=fill, align=wrap)
        ws2.cell(row, 1).alignment = center
        ws2.cell(row, 3).alignment = center
        ws2.cell(row, 4).alignment = center

    total_unique_obj = len({d['obj'] for d in rows_data if d['obj']})
    for label, val1, val2 in [('ИТОГО:', total_eng, ''), ('Итого объектов:', '', total_unique_obj)]:
        ws2.append(['', label, val1, val2])
        row = ws2.max_row
        for cell in ws2[row]:
            _style(cell, font=hdr_font, fill=tot_fill, align=center)

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14

    return wb


def generate_rasstanovka(pril7_path, template_path, output_dir, report_date='', fmt='html', log_func=print):
    """Возвращает (output_path, table_data). fmt='html' или 'excel'."""
    rows_data, today = _prepare_data(pril7_path, template_path, report_date, log_func)
    if rows_data is None:
        return None, []

    # Сортируем по порядку должностей, внутри — по ФИО
    rows_data.sort(key=lambda d: (_dolzh_sort_key(d['dolzh']), d['fio']))

    # Группировка по должности (после сортировки одинаковые должности стоят рядом)
    groups = []
    cur_dolzh, cur_group = None, []
    for d in rows_data:
        if d['dolzh'] != cur_dolzh:
            if cur_group:
                groups.append((cur_dolzh, cur_group))
            cur_dolzh, cur_group = d['dolzh'], []
        cur_group.append(d)
    if cur_group:
        groups.append((cur_dolzh, cur_group))

    if fmt == 'excel':
        out_name = f'Отчёт-расстановка {today}.xlsx'
        out_path = os.path.join(output_dir, out_name)
        try:
            wb = _build_excel(rows_data, today, groups)
            wb.save(out_path)
        except OSError as e:
            log_func(f'ОШИБКА записи файла: {e}')
            return None, []
    else:
        out_name = f'Отчёт-расстановка {today}.html'
        out_path = os.path.join(output_dir, out_name)
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(_build_html(rows_data, today, groups))
        except OSError as e:
            log_func(f'ОШИБКА записи файла: {e}')
            return None, []

    table_data = [[i+1, d['obj'][:60], d['contractor'][:40], d['dolzh'][:40], d['fio'], d['tel']]
                  for i, d in enumerate(rows_data)]
    log_func(f'Готово! Строк: {len(rows_data)}. Файл: {out_name}')
    return out_path, table_data
