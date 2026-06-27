"""Заполнение Приложения 7 (openpyxl, keep_vba для .xlsm)."""

from __future__ import annotations

import traceback
from datetime import datetime
from typing import Callable

from sk_reporter.deployment.lookup import DEFAULT_DESC, load_desc_map
from sk_reporter.deployment.summary import read_summary_rows
from sk_reporter.deployment.xlsx_save import save_xlsm_workbook

SHEET_NAME = "Отчет о ВОУ"
DATE_ROW = 12
FIRST_COL = 6
LAST_COL = 36
SUM_COL = 37


def _prepare_df(summary_path, log_func):
    log_func("Читаю summary...")
    try:
        rows = read_summary_rows(summary_path)
    except Exception as exc:
        log_func(f"ОШИБКА: {exc}")
        return None
    if not rows:
        log_func("ОШИБКА: summary пуст")
        return None
    # emulate pandas-like row access
    class Row:
        __slots__ = ("_d",)
        def __init__(self, d): self._d = d
        def __getitem__(self, k): return self._d[k]
    filtered = []
    for r in rows:
        if not r.get("Инженер СК") or not r.get("Дата") or not r.get("Объект"):
            continue
        r = dict(r)
        r["Инженер СК"] = r["Инженер СК"].strip()
        r["Дата"] = str(r["Дата"]).strip()
        r["Объект"] = r["Объект"].strip()
        r["Генподрядчик"] = (r.get("Генподрядчик") or "").strip()
        filtered.append(Row(r))
    if not filtered:
        log_func("ОШИБКА: нет валидных строк в summary")
        return None
    return filtered


def _build_groups(df):
    groups = {}
    for row in df:
        key = (row["Инженер СК"], row["Дата"])
        groups.setdefault(key, []).append(row)
    return groups


def _shares(n):
    if n <= 0:
        return []
    base = 100 // n
    remainder = 100 - base * n
    vals = [base / 100] * n
    if remainder:
        vals[-1] = (base + remainder) / 100
    return vals


def _fill_openpyxl(df, pril7_path, log_func, desc_map=None):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from copy import copy

    def col(n): return get_column_letter(n)

    def _read(ws, r, c):
        cell = ws.cell(r, c)
        try:
            return cell.value
        except Exception:
            return None

    def _write(ws, r, c, v):
        try:
            ws.cell(r, c).value = v
        except AttributeError:
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                    try: ws.cell(rng.min_row, rng.min_col).value = v
                    except AttributeError: pass
                    return

    def _s(ws, r, c):
        v = _read(ws, r, c)
        return (str(v).strip() if v is not None else "")

    def _find_date_col(ws, date_str):
        try:
            t = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            return None
        for c in range(FIRST_COL, LAST_COL + 1):
            v = _read(ws, DATE_ROW, c)
            if isinstance(v, datetime) and v.date() == t:
                return c
        return None

    def _find_eng(ws, fio, mx):
        for r in range(13, mx + 1):
            if _s(ws, r, 3) == fio and _s(ws, r, 2) == "" and _s(ws, r, 5) == "":
                return r
        return None

    def _last_obj(ws, eng, fio, mx):
        last = eng
        for r in range(eng + 1, mx + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if e == fio: last = r
        return last

    def _find_obj(ws, eng, fio, obj, podr, mx):
        for r in range(eng + 1, mx + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if e == fio and c3 == obj and b == podr: return r
        return None

    def _copy_fmt(ws, src, dst):
        for c in range(1, ws.max_column + 1):
            try:
                sc = ws.cell(src, c); dc = ws.cell(dst, c)
                if sc.has_style:
                    dc._style = copy(sc._style)
            except AttributeError:
                pass
        sd = ws.row_dimensions.get(src)
        if sd:
            ws.row_dimensions[dst].height       = sd.height
            ws.row_dimensions[dst].outline_level = sd.outline_level

    def _sibling_obj(ws, eng):
        """Первая строка объекта внутри блока eng — для копирования стиля."""
        for r in range(eng + 1, ws.max_row + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if b or e: return r
        return None

    def _tmpl_eng(ws):
        for r in range(13, ws.max_row + 1):
            if _s(ws, r, 3) and not _s(ws, r, 2) and not _s(ws, r, 5): return r
        return None

    def _clear_date_col(ws, eng, fio, date_col):
        for r in range(eng + 1, ws.max_row + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e:  # следующий заголовок — стоп
                break
            if b and e:                 # только строки с подрядчиком и ФИО
                try: ws.cell(r, date_col).value = None
                except AttributeError: pass

    def _upd_sum(ws, eng, f_obj, l_obj):
        for c in range(FIRST_COL, LAST_COL + 1):
            _write(ws, eng, c, f"=SUM({col(c)}{f_obj}:{col(c)}{l_obj})")
        _write(ws, eng, SUM_COL,
               f"=SUM({col(FIRST_COL)}{eng}:{col(LAST_COL)}{eng})")

    def _add_eng(ws, fio):
        last = ws.max_row
        while last > 13 and not _read(ws, last, 3): last -= 1
        nr = last + 2
        t = _tmpl_eng(ws)
        if t:
            _copy_fmt(ws, t, nr)
            for c in range(1, ws.max_column + 1):
                try: ws.cell(nr, c).value = None
                except AttributeError: pass
        _write(ws, nr, 3, fio)
        ws.row_dimensions[nr].outline_level = 0
        ws.row_dimensions[nr].hidden = False
        return nr

    def _add_obj(ws, eng, fio, podr, obj):
        lo = _last_obj(ws, eng, fio, ws.max_row)
        nr = lo + 1
        ws.insert_rows(nr)
        # Шаблон: первая существующая строка объекта в том же блоке
        src = _sibling_obj(ws, eng)
        if src:
            _copy_fmt(ws, src, nr)
            for c in range(1, ws.max_column + 1):
                try: ws.cell(nr, c).value = None
                except AttributeError: pass
        _write(ws, nr, 2, podr)
        _write(ws, nr, 3, obj)
        _write(ws, nr, 4, (desc_map or {}).get(fio, DEFAULT_DESC))
        _write(ws, nr, 5, fio)
        _write(ws, nr, SUM_COL,
               f"=SUM({col(FIRST_COL)}{nr}:{col(LAST_COL)}{nr})")
        ws.row_dimensions[nr].outline_level = 1
        ws.row_dimensions[nr].hidden = False
        lo2 = _last_obj(ws, eng, fio, ws.max_row)
        _upd_sum(ws, eng, eng + 1, lo2)
        return nr

    log_func("Открываю Приложение 7 (openpyxl)...")
    keep_vba = str(pril7_path).lower().endswith('.xlsm')
    wb = load_workbook(pril7_path, keep_vba=keep_vba)
    if SHEET_NAME not in wb.sheetnames:
        log_func(f"ОШИБКА: Лист '{SHEET_NAME}' не найден.")
        return False
    ws = wb[SHEET_NAME]
    processed = 0

    for (fio, date), rows in _build_groups(df).items():
        try:
            dc = _find_date_col(ws, date)
            if dc is None:
                log_func(f"  Дата {date} не найдена — пропуск")
                continue
            eng = _find_eng(ws, fio, ws.max_row)
            if eng is None:
                log_func(f"  Новый инженер: {fio}")
                eng = _add_eng(ws, fio)
            _clear_date_col(ws, eng, fio, dc)
            vals = _shares(len(rows))
            for i, row in enumerate(rows):
                obj  = row["Объект"]
                podr = row["Генподрядчик"]
                obj_r = _find_obj(ws, eng, fio, obj, podr, ws.max_row)
                if obj_r is None:
                    log_func(f"  Новый объект: {obj[:60]}")
                    obj_r = _add_obj(ws, eng, fio, podr, obj)
                _write(ws, obj_r, 4, (desc_map or {}).get(fio, DEFAULT_DESC))
                _write(ws, obj_r, dc, vals[i])
                processed += 1
        except Exception as e:
            log_func(f"  ОШИБКА [{fio} | {date}]: {e}")
            log_func(traceback.format_exc())

    log_func("Сохраняю файл...")
    try:
        save_xlsm_workbook(wb, pril7_path)
    except Exception as e:
        log_func(f"ОШИБКА при сохранении: {e}")
        log_func(traceback.format_exc())
        return False

    log_func(f"Готово! Обработано записей: {processed}")
    return True




def fill_pril7(summary_path, pril7_path, log_func: Callable[[str], None] = print) -> bool:
    df = _prepare_df(summary_path, log_func)
    if df is None:
        return False
    desc_map = load_desc_map()
    log_func(f"Справочник описаний: {len(desc_map)} записей")
    return _fill_openpyxl(df, pril7_path, log_func, desc_map)
