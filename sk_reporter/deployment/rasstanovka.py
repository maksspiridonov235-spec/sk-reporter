"""Отчёт-расстановка из заполненного Прил.7."""

from __future__ import annotations

import traceback
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Callable

import openpyxl

from sk_reporter.deployment.lookup import DEFAULT_REZHIM, load_sprav_dict

SHEET_RAB = "Рабочая исходник"
SHEET_REPORT = "Отчет расстановка"
DATA_START_ROW = 5

_DOLZH_ORDER = [
    "Инженер СК (общестроительные работы, сварочные технологии)",
    "Инженер СК (общестроительные работы)",
    "Инженер СК (электромонтажные работы, КИПиА)",
    "Инженер ПИЛ  (НК, УЗК)",
    "Инженер СЛ НК",
    "Инженер ОЗОТОБОС СПД",
    "Инженер ОЗОТОБОС",
    "Инженер БДД",
    "Инженер СК (Геодезические работы)",
    "Инженер СК (Геологоизыскательские работы)",
    "Системотехник / Инженер ПТО",
    "Инженер ПТО СПД СГМ",
    "Инженер ПТО СГМ",
    "Инженер ПТО Тюмень",
    "Руководитель СК",
]
_DOLZH_IDX = {d: i for i, d in enumerate(_DOLZH_ORDER)}


def _read_rabochaya(template_path):
    """Читает Рабочая исходник → {ФИО: {zakr_fio, zakr_tel}}."""
    if not template_path or not Path(template_path).exists():
        return {}
    try:
        wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=True)
        if SHEET_RAB not in wb.sheetnames:
            return {}
        ws = wb[SHEET_RAB]
        result = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            fio = row[2]
            if not fio or not isinstance(fio, str):
                continue
            fio = fio.strip()
            if not fio or fio == 'ФИО':
                continue
            result[fio] = {
                'zakr_fio': str(row[6]).strip() if row[6] else '',
                'zakr_tel': str(row[7]).strip() if row[7] else '',
            }
        return result
    except Exception:
        return {}


def _parse_excel_date(v):
    """Пробует преобразовать значение ячейки Excel в date. None если не дата."""
    if v is None:
        return None
    try:
        if hasattr(v, 'date'):
            return v.date()
        if isinstance(v, (int, float)) and v > 20000:
            return (datetime(1899, 12, 30) + __import__('datetime').timedelta(days=int(v))).date()
    except Exception:
        pass
    return None


def _is_date_like(v):
    """Грубая проверка — похоже ли значение на дату/число дня."""
    if v is None:
        return False
    if hasattr(v, 'date'):
        return True
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv > 20000 or (1 <= iv <= 31)
    if isinstance(v, str):
        s = v.strip().replace('.', '').replace(',', '')
        if s.isdigit():
            return 1 <= int(s) <= 31
    return False


def _read_pril7_openpyxl(pril7_path, log_func, override_date=None):
    """Читает Приложение 7 через openpyxl — без COM, безопасно в потоках Flask."""
    rows = []
    try:
        wb = openpyxl.load_workbook(pril7_path, keep_vba=True, data_only=True)
        if 'Отчет о ВОУ' not in wb.sheetnames:
            log_func('  ОШИБКА: лист "Отчет о ВОУ" не найден')
            return rows
        ws = wb['Отчет о ВОУ']

        # 1. Дата отчёта
        if override_date:
            report_date = override_date
            log_func(f'  Дата (из интерфейса): {report_date.strftime("%d.%m.%Y")}')
        else:
            report_date = _parse_excel_date(ws.cell(4, 48).value)
            if report_date:
                log_func(f'  Дата (из AV4): {report_date.strftime("%d.%m.%Y")}')
            else:
                log_func('  ПРЕДУПРЕЖДЕНИЕ: дата в AV4 не найдена — берём все строки с ФИО и Объектом')

        # 2. Строка заголовков с датами (строки 7–16)
        header_row, max_count = 0, 0
        for r in range(7, 17):
            cnt = sum(1 for c in range(6, 61) if _is_date_like(ws.cell(r, c).value))
            if cnt > max_count:
                max_count, header_row = cnt, r
        if header_row == 0:
            header_row = 12
        log_func(f'  Строка дат: {header_row}')

        # 3. Столбец нужной даты
        today_col = None
        if report_date:
            for c in range(6, 61):
                d = _parse_excel_date(ws.cell(header_row, c).value)
                if d and d == report_date:
                    today_col = c
                    break
                v = ws.cell(header_row, c).value
                if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
                    if int(v) == report_date.day:
                        today_col = c
                        break
            if today_col:
                log_func(f'  Столбец даты: {today_col}')
            else:
                log_func('  Столбец даты не найден — берём все строки')

        # 4. Сбор данных
        data_start = header_row + 1
        max_row = ws.max_row
        log_func(f'  Читаю строки {data_start}–{max_row}')

        for r in range(data_start, max_row + 1):
            fio_val = ws.cell(r, 5).value   # E
            obj_val = ws.cell(r, 3).value   # C
            if not fio_val or not obj_val:
                continue
            if today_col and not ws.cell(r, today_col).value:
                continue
            rows.append({
                'fio':        str(fio_val).strip(),
                'obj':        str(obj_val).strip(),
                'contractor': str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else '',
            })

        log_func(f'  Найдено строк: {len(rows)}')
        wb.close()

    except Exception as e:
        log_func(f'  ОШИБКА чтения Приложения 7: {e}')
        import traceback as tb
        log_func(tb.format_exc())
    return rows


def _prepare_data(pril7_path, template_path, report_date, log_func):
    """Общая подготовка данных. Возвращает (rows_data, today_str) или (None, None)."""
    override_date = None
    if report_date:
        try:
            override_date = datetime.strptime(report_date, '%Y-%m-%d').date()
            log_func(f'Дата из интерфейса: {override_date.strftime("%d.%m.%Y")}')
        except ValueError:
            log_func(f'Неверный формат даты: {report_date}, читаю из AV4')

    log_func('Читаю справочник...')
    sprav = load_sprav_dict()

    log_func('Читаю Рабочая исходник из шаблона...')
    rab_map = _read_rabochaya(template_path)

    log_func('Читаю Приложение 7...')
    pril7_rows = _read_pril7_openpyxl(pril7_path, log_func, override_date=override_date)

    if not pril7_rows:
        log_func('ОШИБКА: нет данных в Приложении 7')
        return None, None

    log_func(f'Найдено строк: {len(pril7_rows)}')
    today = override_date.strftime('%d.%m.%Y') if override_date else datetime.now().strftime('%d.%m.%Y')

    rows_data = []
    for item in pril7_rows:
        fio  = item['fio']
        sp   = sprav.get(fio, {})
        rb   = rab_map.get(fio, {})
        rows_data.append({
            'obj':        item['obj'],
            'rezhim':     sp.get('rezhim', '').strip() or DEFAULT_REZHIM,
            'contractor': item['contractor'],
            'dolzh':      sp.get('dolzhnost', ''),
            'fio':        fio,
            'tel':        sp.get('telefon', ''),
            'zakr_fio':   rb.get('zakr_fio', ''),
            'zakr_tel':   rb.get('zakr_tel', ''),
        })
    return rows_data, today


def _build_excel(rows_data, today, groups):
    """Создаёт openpyxl Workbook с листами Расстановка и Итоги."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_font  = Font(bold=True)
    hdr_fill  = PatternFill('solid', fgColor='C6D9F1')
    grp_fill  = PatternFill('solid', fgColor='DCE6F1')
    tot_fill  = PatternFill('solid', fgColor='BDD7EE')
    alt_fill  = PatternFill('solid', fgColor='F0F6FF')
    thin      = Side(style='thin', color='999999')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='top', wrap_text=True)
    wrap      = Alignment(vertical='top', wrap_text=True)

    def _style(cell, font=None, fill=None, align=None):
        cell.border = border
        if font:  cell.font  = font
        if fill:  cell.fill  = fill
        if align: cell.alignment = align

    # ── Лист 1: Расстановка ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Расстановка'

    # Строка 1 — заголовок документа
    ws1.append([f'Отчёт-расстановка на {today}'])
    ws1.merge_cells('A1:H1')
    title_cell = ws1.cell(1, 1)
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 22

    # Строка 2 — шапка таблицы
    headers = [
        '№ п/п',
        'Наименование объектов',
        'Режим контроля',
        'Наименование подрядной организации',
        'Должность СК',
        'ФИО закреплённого специалиста СК',
        'Телефон',
        'Закреплённое за специалистом строительного контроля, ФИО водителя, телефон',
    ]
    ws1.append(headers)
    for cell in ws1[2]:
        _style(cell, font=hdr_font, fill=hdr_fill, align=center)

    num = 0
    for _, group in groups:
        for d in group:
            num += 1
            zakr = ' '.join(filter(None, [d['zakr_fio'], d['zakr_tel']]))
            ws1.append([num, d['obj'], d['rezhim'], d['contractor'],
                        d['dolzh'], d['fio'], d['tel'], zakr])
            row = ws1.max_row
            fill = alt_fill if num % 2 else None
            for cell in ws1[row]:
                _style(cell, fill=fill, align=wrap)
            ws1.cell(row, 1).alignment = center

    # Строки итогов
    total_eng_main = len({d['fio'] for d in rows_data})
    total_obj_main = len({d['obj'] for d in rows_data if d['obj']})
    for label, val in [('Итого персонала:', total_eng_main),
                        ('Итого объектов:',  total_obj_main)]:
        ws1.append(['', '', '', '', label, val, '', ''])
        row = ws1.max_row
        ws1.merge_cells(f'A{row}:D{row}')
        ws1.merge_cells(f'F{row}:H{row}')
        for cell in ws1[row]:
            _style(cell, font=hdr_font, fill=tot_fill)
        ws1.cell(row, 5).alignment = Alignment(horizontal='right',  vertical='center')
        ws1.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')

    col_widths = [5, 35, 28, 35, 35, 30, 15, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A3'

    # ── Лист 2: Итоги ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Итоги')
    ws2.append(['№', 'Должность', 'Инженеров', 'Объектов'])
    for cell in ws2[1]:
        _style(cell, font=hdr_font, fill=hdr_fill, align=center)

    total_eng = 0
    for i, (dolzh, group) in enumerate(groups):
        grp_eng = len({d['fio'] for d in group})
        grp_obj = len({d['obj'] for d in group if d['obj']})
        total_eng += grp_eng
        ws2.append([i + 1, dolzh or '—', grp_eng, grp_obj])
        row = ws2.max_row
        fill = alt_fill if i % 2 else None
        for cell in ws2[row]:
            _style(cell, fill=fill, align=wrap)
        ws2.cell(row, 1).alignment = center
        ws2.cell(row, 3).alignment = center
        ws2.cell(row, 4).alignment = center

    total_unique_obj = len({d['obj'] for d in rows_data if d['obj']})
    for label, val1, val2 in [('ИТОГО:', total_eng, ''), ('Итого объектов:', '', total_unique_obj)]:
        ws2.append(['', label, val1, val2])
        row = ws2.max_row
        for cell in ws2[row]:
            _style(cell, font=hdr_font, fill=tot_fill, align=center)

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14

    return wb


def _fill_template_report(
    template_path: str | Path,
    out_path: Path,
    rows_data: list[dict],
    log_func: Callable[[str], None],
) -> bool:
    """Копирует шаблон и заполняет лист «Отчет расстановка»."""
    import shutil

    template_path = Path(template_path)
    if not template_path.is_file():
        return False
    shutil.copy2(template_path, out_path)
    try:
        wb = openpyxl.load_workbook(out_path, keep_vba=True)
    except Exception as exc:
        log_func(f"  ОШИБКА открытия шаблона: {exc}")
        return False
    if SHEET_REPORT not in wb.sheetnames:
        log_func(f"  Лист «{SHEET_REPORT}» не найден — простой Excel")
        wb.close()
        return False

    ws = wb[SHEET_REPORT]
    total_obj_row = total_emp_row = None
    for r in range(ws.max_row, DATA_START_ROW - 1, -1):
        label = ws.cell(r, 2).value
        if label == "Итого объектов":
            total_obj_row = r
        elif label == "Количество сотрудников":
            total_emp_row = r

    end_clear = (min(total_obj_row, total_emp_row) - 1) if total_obj_row and total_emp_row else ws.max_row
    for r in range(DATA_START_ROW, end_clear + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None

    write_row = DATA_START_ROW
    for num, d in enumerate(rows_data, 1):
        zakr = " ".join(filter(None, [d["zakr_fio"], d["zakr_tel"]]))
        ws.cell(write_row, 1, num)
        ws.cell(write_row, 2, d["obj"])
        ws.cell(write_row, 3, d["rezhim"])
        ws.cell(write_row, 4, d["contractor"] or None)
        ws.cell(write_row, 5, d["dolzh"])
        ws.cell(write_row, 6, d["fio"])
        ws.cell(write_row, 7, d["tel"] or None)
        ws.cell(write_row, 8, zakr or None)
        write_row += 1

    if total_obj_row:
        ws.cell(total_obj_row, 3, len({d["obj"] for d in rows_data if d["obj"]}))
    if total_emp_row:
        ws.cell(total_emp_row, 3, len({d["fio"] for d in rows_data}))

    try:
        wb.save(out_path)
    except OSError as exc:
        log_func(f"  ОШИБКА сохранения: {exc}")
        return False
    finally:
        wb.close()
    log_func(f"  Заполнен лист «{SHEET_REPORT}» по шаблону {template_path.name}")
    return True


def generate_rasstanovka(
    pril7_path: str | Path,
    template_path: str | Path,
    output_dir: str | Path,
    report_date: str = "",
    log_func: Callable[[str], None] = print,
) -> Path | None:
    rows_data, today = _prepare_data(str(pril7_path), str(template_path), report_date, log_func)
    if rows_data is None:
        return None

    rows_data.sort(key=lambda d: (_DOLZH_IDX.get(d["dolzh"], 999), d["fio"]))

    groups = []
    cur_dolzh, cur_group = None, []
    for d in rows_data:
        if d["dolzh"] != cur_dolzh:
            if cur_group:
                groups.append((cur_dolzh, cur_group))
            cur_dolzh, cur_group = d["dolzh"], []
        cur_group.append(d)
    if cur_group:
        groups.append((cur_dolzh, cur_group))

    out_name = f"Отчёт-расстановка {today}.xlsm"
    out_path = Path(output_dir) / out_name
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    try:
        if _fill_template_report(template_path, out_path, rows_data, log_func):
            log_func(f"Готово! Строк: {len(rows_data)}. Файл: {out_name}")
            return out_path
        log_func("  Шаблон не применён — формирую упрощённый Excel")
        wb = _build_excel(rows_data, today, groups)
        wb.save(out_path)
    except OSError as exc:
        log_func(f"ОШИБКА записи файла: {exc}")
        return None

    log_func(f"Готово! Строк: {len(rows_data)}. Файл: {out_name}")
    return out_path
