"""
fill_pril7.py
  - .xlsx  → openpyxl
  - .xlsm  → win32com (Excel COM) — обходит ограничения openpyxl с rich text
"""
from datetime import datetime
import os
import sys
import traceback
from pathlib import Path

import pandas as pd

from sk_reporter.deployment.lookup import DEFAULT_DESC as _DEFAULT_DESC, load_desc_map as _load_desc_map

SHEET_NAME = "Отчет о ВОУ"
DATE_ROW   = 12
FIRST_COL  = 6    # F
LAST_COL   = 36   # AJ
SUM_COL    = 37   # AK

# ─────────────────────────────────────────────────────────────────────────────
#  Общая подготовка данных
# ─────────────────────────────────────────────────────────────────────────────
def _prepare_df(summary_path, log_func):
    log_func("Читаю исходник...")
    df = pd.read_excel(summary_path)

    required = {"Инженер СК", "Дата", "Объект", "Генподрядчик"}
    if not required.issubset(df.columns):
        log_func(f"ОШИБКА: Нет колонок: {required - set(df.columns)}")
        return None

    df = df.dropna(subset=["Инженер СК", "Дата", "Объект"])
    df["Инженер СК"]   = df["Инженер СК"].str.strip()
    df["Дата"]         = df["Дата"].astype(str).str.strip()
    df["Объект"]       = df["Объект"].str.strip()
    df["Генподрядчик"] = df["Генподрядчик"].fillna("").str.strip()
    return df


def _build_groups(df):
    """Возвращает {(fio, date): [row, ...]} — сгруппировано для расчёта долей."""
    groups: dict = {}
    for _, row in df.iterrows():
        key = (row["Инженер СК"], row["Дата"])
        groups.setdefault(key, []).append(row)
    return groups


def _shares(n):
    """Возвращает n долей с точностью до 0.01, сумма строго равна 1.0.
    Остаток от деления уходит в последний элемент."""
    if n <= 0:
        return []
    base = 100 // n          # целые центы
    remainder = 100 - base * n
    vals = [base / 100] * n
    if remainder:
        vals[-1] = (base + remainder) / 100
    return vals


# ═════════════════════════════════════════════════════════════════════════════
#  РЕЖИМ 1: openpyxl  (.xlsx)
# ═════════════════════════════════════════════════════════════════════════════
def _fill_openpyxl(df, pril7_path, log_func, desc_map=None, *, save_as_xlsm=False):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from copy import copy

    def col(n): return get_column_letter(n)

    def _read(ws, r, c):
        cell = ws.cell(r, c)
        try:
            return cell.value
        except Exception:
            return None

    def _write(ws, r, c, v):
        try:
            ws.cell(r, c).value = v
        except AttributeError:
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                    try: ws.cell(rng.min_row, rng.min_col).value = v
                    except AttributeError: pass
                    return

    def _s(ws, r, c):
        v = _read(ws, r, c)
        return (str(v).strip() if v is not None else "")

    def _find_date_col(ws, date_str):
        try:
            t = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            return None
        for c in range(FIRST_COL, LAST_COL + 1):
            v = _read(ws, DATE_ROW, c)
            if isinstance(v, datetime) and v.date() == t:
                return c
        return None

    def _find_eng(ws, fio, mx):
        for r in range(13, mx + 1):
            if _s(ws, r, 3) == fio and _s(ws, r, 2) == "" and _s(ws, r, 5) == "":
                return r
        return None

    def _last_obj(ws, eng, fio, mx):
        last = eng
        for r in range(eng + 1, mx + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if e == fio: last = r
        return last

    def _find_obj(ws, eng, fio, obj, podr, mx):
        for r in range(eng + 1, mx + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if e == fio and c3 == obj and b == podr: return r
        return None

    def _copy_fmt(ws, src, dst):
        for c in range(1, ws.max_column + 1):
            try:
                sc = ws.cell(src, c); dc = ws.cell(dst, c)
                if sc.has_style:
                    dc._style = copy(sc._style)
            except AttributeError:
                pass
        sd = ws.row_dimensions.get(src)
        if sd:
            ws.row_dimensions[dst].height       = sd.height
            ws.row_dimensions[dst].outline_level = sd.outline_level

    def _sibling_obj(ws, eng):
        """Первая строка объекта внутри блока eng — для копирования стиля."""
        for r in range(eng + 1, ws.max_row + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e: break
            if b or e: return r
        return None

    def _tmpl_eng(ws):
        for r in range(13, ws.max_row + 1):
            if _s(ws, r, 3) and not _s(ws, r, 2) and not _s(ws, r, 5): return r
        return None

    def _clear_date_col(ws, eng, fio, date_col):
        for r in range(eng + 1, ws.max_row + 1):
            c3 = _s(ws, r, 3); b = _s(ws, r, 2); e = _s(ws, r, 5)
            if c3 and not b and not e:  # следующий заголовок — стоп
                break
            if b and e:                 # только строки с подрядчиком и ФИО
                try: ws.cell(r, date_col).value = None
                except AttributeError: pass

    def _upd_sum(ws, eng, f_obj, l_obj):
        for c in range(FIRST_COL, LAST_COL + 1):
            _write(ws, eng, c, f"=SUM({col(c)}{f_obj}:{col(c)}{l_obj})")
        _write(ws, eng, SUM_COL,
               f"=SUM({col(FIRST_COL)}{eng}:{col(LAST_COL)}{eng})")

    def _add_eng(ws, fio):
        last = ws.max_row
        while last > 13 and not _read(ws, last, 3): last -= 1
        nr = last + 2
        ws.insert_rows(nr)
        t = _tmpl_eng(ws)
        if t:
            _copy_fmt(ws, t, nr)
            for c in range(1, ws.max_column + 1):
                try: ws.cell(nr, c).value = None
                except AttributeError: pass
        _write(ws, nr, 3, fio)
        ws.row_dimensions[nr].outline_level = 0
        ws.row_dimensions[nr].hidden = False
        return nr

    def _add_obj(ws, eng, fio, podr, obj):
        lo = _last_obj(ws, eng, fio, ws.max_row)
        nr = lo + 1
        ws.insert_rows(nr)
        # Шаблон: первая существующая строка объекта в том же блоке
        src = _sibling_obj(ws, eng)
        if src:
            _copy_fmt(ws, src, nr)
            for c in range(1, ws.max_column + 1):
                try: ws.cell(nr, c).value = None
                except AttributeError: pass
        _write(ws, nr, 2, podr)
        _write(ws, nr, 3, obj)
        _write(ws, nr, 4, (desc_map or {}).get(fio, _DEFAULT_DESC))
        _write(ws, nr, 5, fio)
        _write(ws, nr, SUM_COL,
               f"=SUM({col(FIRST_COL)}{nr}:{col(LAST_COL)}{nr})")
        ws.row_dimensions[nr].outline_level = 1
        ws.row_dimensions[nr].hidden = False
        lo2 = _last_obj(ws, eng, fio, ws.max_row)
        _upd_sum(ws, eng, eng + 1, lo2)
        return nr

    log_func("Открываю Приложение 7 (openpyxl)...")
    wb = load_workbook(pril7_path, keep_vba=save_as_xlsm)
    if SHEET_NAME not in wb.sheetnames:
        log_func(f"ОШИБКА: Лист '{SHEET_NAME}' не найден.")
        return False
    ws = wb[SHEET_NAME]
    processed = 0

    for (fio, date), rows in _build_groups(df).items():
        try:
            dc = _find_date_col(ws, date)
            if dc is None:
                log_func(f"  Дата {date} не найдена — пропуск")
                continue
            eng = _find_eng(ws, fio, ws.max_row)
            if eng is None:
                log_func(f"  Новый инженер: {fio}")
                eng = _add_eng(ws, fio)
            _clear_date_col(ws, eng, fio, dc)
            vals = _shares(len(rows))
            for i, row in enumerate(rows):
                obj  = row["Объект"]
                podr = row["Генподрядчик"]
                obj_r = _find_obj(ws, eng, fio, obj, podr, ws.max_row)
                if obj_r is None:
                    log_func(f"  Новый объект: {obj[:60]}")
                    obj_r = _add_obj(ws, eng, fio, podr, obj)
                _write(ws, obj_r, 4, (desc_map or {}).get(fio, _DEFAULT_DESC))
                _write(ws, obj_r, dc, vals[i])
                processed += 1
        except Exception as e:
            log_func(f"  ОШИБКА [{fio} | {date}]: {e}")
            log_func(traceback.format_exc())

    log_func("Сохраняю файл...")
    try:
        if save_as_xlsm:
            from sk_reporter.deployment.xlsx_save import save_xlsm_workbook
            save_xlsm_workbook(wb, pril7_path)
        else:
            wb.save(pril7_path)
    except Exception as e:
        log_func(f"ОШИБКА при сохранении: {e}")
        log_func(traceback.format_exc())
        return False

    log_func(f"Готово! Обработано записей: {processed}")
    return True


# ═════════════════════════════════════════════════════════════════════════════
#  РЕЖИМ 2: win32com  (.xlsm)
# ═════════════════════════════════════════════════════════════════════════════
def _fill_win32(df, pril7_path, log_func, desc_map=None):
    try:
        import win32com.client as win32
        import pythoncom
    except ImportError:
        log_func("ОШИБКА: win32com не установлен. Запустите: pip install pywin32")
        return False

    XL_WHOLE  = 2
    XL_DOWN   = -4121
    XL_LEFT   = -4159

    abs_path = os.path.abspath(pril7_path)
    log_func("Открываю Приложение 7 через Excel COM...")

    pythoncom.CoInitialize()

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(abs_path)
        ws = wb.Sheets(SHEET_NAME)

        def _cell(r, c): return ws.Cells(r, c)
        def _val(r, c):
            v = _cell(r, c).Value
            return str(v).strip() if v is not None else ""

        def _find_date_col(date_str):
            try: target = datetime.strptime(date_str, "%d.%m.%Y").date()
            except ValueError: return None
            for c in range(FIRST_COL, LAST_COL + 1):
                v = _cell(DATE_ROW, c).Value
                if isinstance(v, datetime) and v.date() == target: return c
            return None

        def _max_row():
            return ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1

        def _find_eng(fio):
            mx = _max_row()
            for r in range(13, mx + 1):
                if _val(r, 3) == fio and _val(r, 2) == "" and _val(r, 5) == "":
                    return r
            return None

        def _last_obj_row(eng, fio):
            mx = _max_row(); last = eng
            for r in range(eng + 1, mx + 1):
                c3 = _val(r, 3); b = _val(r, 2); e = _val(r, 5)
                if c3 and not b and not e: break
                if e == fio: last = r
            return last

        def _find_obj(eng, fio, obj, podr):
            mx = _max_row()
            for r in range(eng + 1, mx + 1):
                c3 = _val(r, 3); b = _val(r, 2); e = _val(r, 5)
                if c3 and not b and not e: break
                if e == fio and c3 == obj and b == podr: return r
            return None

        def _upd_sum(eng, f_obj, l_obj):
            for c in range(FIRST_COL, LAST_COL + 1):
                cl = chr(64 + c) if c <= 26 else chr(64 + c // 26) + chr(64 + c % 26)
                _cell(eng, c).Formula = f"=SUM({cl}{f_obj}:{cl}{l_obj})"

        def _clear_date_col(eng, fio, date_col):
            mx = _max_row()
            for r in range(eng + 1, mx + 1):
                c3 = _val(r, 3); b = _val(r, 2); e = _val(r, 5)
                if c3 and not b and not e:  # следующий заголовок — стоп
                    break
                if b and e:                 # только строки с подрядчиком и ФИО
                    try: _cell(r, date_col).ClearContents()
                    except Exception: pass

        def _col_letter(c):
            if c <= 26: return chr(64 + c)
            return chr(64 + (c - 1) // 26) + chr(64 + (c - 1) % 26 + 1)

        def _add_eng(fio):
            mx = _max_row()
            nr = mx + 2
            # Скопировать формат из существующего заголовка
            for r in range(13, mx + 1):
                if _val(r, 3) and not _val(r, 2) and not _val(r, 5):
                    ws.Rows(r).Copy()
                    ws.Rows(nr).Insert()
                    # Очистить значения в новой строке
                    for c in range(1, LAST_COL + 5):
                        try: _cell(nr, c).ClearContents()
                        except: pass
                    break
            _cell(nr, 3).Value = fio
            ws.Rows(nr).OutlineLevel = 1
            return nr

        def _sibling_obj_com(eng):
            """Первая строка объекта внутри блока eng."""
            mx = _max_row()
            for r in range(eng + 1, mx + 1):
                c3 = _val(r, 3); b = _val(r, 2); e = _val(r, 5)
                if c3 and not b and not e: break
                if b or e: return r
            return None

        def _copy_row_style_com(src, dst):
            """Копирует стиль ячеек строки src в dst (без значений)."""
            ws.Rows(dst).RowHeight = ws.Rows(src).RowHeight
            for c in range(1, LAST_COL + 6):
                try:
                    sc = _cell(src, c); dc = _cell(dst, c)
                    dc.Interior.Color      = sc.Interior.Color
                    dc.Interior.Pattern    = sc.Interior.Pattern
                    dc.Font.Name           = sc.Font.Name
                    dc.Font.Size           = sc.Font.Size
                    dc.Font.Bold           = sc.Font.Bold
                    dc.Font.Color          = sc.Font.Color
                    dc.HorizontalAlignment = sc.HorizontalAlignment
                    dc.VerticalAlignment   = sc.VerticalAlignment
                    dc.WrapText            = sc.WrapText
                    dc.NumberFormat        = sc.NumberFormat
                    for b in [7, 8, 9, 10]:
                        try:
                            dc.Borders(b).LineStyle = sc.Borders(b).LineStyle
                            dc.Borders(b).Color     = sc.Borders(b).Color
                            dc.Borders(b).Weight    = sc.Borders(b).Weight
                        except Exception: pass
                except Exception: pass

        def _add_obj(eng, fio, podr, obj):
            lo = _last_obj_row(eng, fio)
            nr = lo + 1
            ws.Rows(nr).Insert()
            # Стиль из первой существующей строки объекта того же блока
            tmpl = _sibling_obj_com(eng)
            if tmpl:
                _copy_row_style_com(tmpl, nr)
            # Очистить значения
            for c in range(1, LAST_COL + 5):
                try: _cell(nr, c).ClearContents()
                except: pass
            # Заполнить данные
            _cell(nr, 2).Value = podr
            _cell(nr, 3).Value = obj
            _cell(nr, 4).Value = (desc_map or {}).get(fio, _DEFAULT_DESC)
            _cell(nr, 5).Value = fio
            # SUM формула
            fl = _col_letter(FIRST_COL); ll = _col_letter(LAST_COL)
            _cell(nr, SUM_COL).Formula = f"=SUM({fl}{nr}:{ll}{nr})"
            ws.Rows(nr).OutlineLevel = 2
            # Обновить SUM в заголовке
            lo2 = _last_obj_row(eng, fio)
            _upd_sum(eng, eng + 1, lo2)
            return nr

        processed = 0
        for (fio, date), rows in _build_groups(df).items():
            try:
                dc = _find_date_col(date)
                if dc is None:
                    log_func(f"  Дата {date} не найдена — пропуск")
                    continue
                eng = _find_eng(fio)
                if eng is None:
                    log_func(f"  Новый инженер: {fio}")
                    eng = _add_eng(fio)
                _clear_date_col(eng, fio, dc)
                vals = _shares(len(rows))
                for i, row in enumerate(rows):
                    obj  = row["Объект"]
                    podr = row["Генподрядчик"]
                    obj_r = _find_obj(eng, fio, obj, podr)
                    if obj_r is None:
                        log_func(f"  Новый объект: {obj[:60]}")
                        obj_r = _add_obj(eng, fio, podr, obj)
                    _cell(obj_r, 4).Value = (desc_map or {}).get(fio, _DEFAULT_DESC)
                    _cell(obj_r, dc).Value = vals[i]
                    processed += 1
            except Exception as e:
                log_func(f"  ОШИБКА [{fio} | {date}]: {e}")
                log_func(traceback.format_exc())

        log_func("Сохраняю файл...")
        wb.Save()
        log_func(f"Готово! Обработано записей: {processed}")
        return True

    except Exception as e:
        log_func(f"ОШИБКА: {e}")
        log_func(traceback.format_exc())
        return False
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


# ═════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ═════════════════════════════════════════════════════════════════════════════
def fill_pril7(summary_path, pril7_path, log_func=print):
    df = _prepare_df(summary_path, log_func)
    if df is None:
        return False

    desc_map = _load_desc_map()
    log_func(f"Справочник описаний: {len(desc_map)} записей")

    if pril7_path.lower().endswith('.xlsm') and sys.platform == 'win32':
        log_func("Формат .xlsm → Excel COM (Windows)")
        return _fill_win32(df, pril7_path, log_func, desc_map)
    if pril7_path.lower().endswith('.xlsm'):
        from sk_reporter.deployment.excel_engine import fill_xlsm_via_libreoffice, libreoffice_available

        if libreoffice_available():
            log_func("Формат .xlsm → LibreOffice + openpyxl")
            return fill_xlsm_via_libreoffice(
                Path(pril7_path),
                lambda xlsx: _fill_openpyxl(df, str(xlsx), log_func, desc_map),
                log_func,
            )
        log_func(
            "ПРЕДУПРЕЖДЕНИЕ: LibreOffice не установлен на сервере — "
            "запись через openpyxl (ограниченный режим, без Excel COM)"
        )
        return _fill_openpyxl(df, pril7_path, log_func, desc_map, save_as_xlsm=True)
    return _fill_openpyxl(df, pril7_path, log_func, desc_map)
