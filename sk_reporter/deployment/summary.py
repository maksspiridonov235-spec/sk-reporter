"""Сбор summary.xlsx из распарсенных строк."""

from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

from sk_reporter.deployment.lookup import resolve_person_fio

COLUMNS = ["Файл", "Дата", "Объект", "Инженер СК", "Генподрядчик"]


def write_summary(rows: list[dict[str, str]], output_path: str | Path) -> Path:
    path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def read_summary_rows(summary_path: str | Path) -> list[dict[str, str]]:
    wb = load_workbook(summary_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(rows_iter)]
    idx = {name: header.index(name) for name in COLUMNS if name in header}
    if len(idx) != len(COLUMNS):
        wb.close()
        raise ValueError(f"Нет колонок summary: {set(COLUMNS) - set(idx)}")
    result: list[dict[str, str]] = []
    for raw in rows_iter:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {col: str(raw[idx[col]] or "").strip() for col in COLUMNS}
        result.append(row)
    wb.close()
    return result


def validate_and_filter_rows(
    rows: list[dict[str, str]],
    *,
    known_fio: set[str] | None = None,
    log_func: Callable[[str], None] = print,
) -> list[dict[str, str]]:
    _ = known_fio  # совместимость; используем resolve_person_fio
    accepted: list[dict[str, str]] = []
    for row in rows:
        fio = row.get("Инженер СК", "").strip()
        obj = row.get("Объект", "").strip()
        fname = row.get("Файл", "")
        date = row.get("Дата", "")

        if date.startswith("Ошибка:"):
            log_func(f"  Пропуск {fname}: {date}")
            continue
        if not fio:
            log_func(f"  Пропуск {fname}: нет инженера СК")
            continue
        canonical = resolve_person_fio(fio)
        if not canonical:
            log_func(f"  Пропуск {fname}: инженер «{fio}» не в справочнике сотрудников")
            continue
        if canonical != fio:
            log_func(f"  {fname}: «{fio}» → «{canonical}»")
            row = dict(row)
            row["Инженер СК"] = canonical
        if not obj:
            log_func(f"  Пропуск {fname}: нет объекта")
            continue
        accepted.append(row)
    return accepted
