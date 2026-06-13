"""
Агент проверки предписания (Excel).

Читает лист «Форма заполнения предписания»: B18 (содержание замечания),
B19 (нормативный документ). Сверка с Техэксперт, при недоступности — с открытым интернетом.
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sk_reporter.prescriptions.techexpert_client import (
    _short_doc_title,
    _title_from_reference,
    lookup_normative,
    parse_normative_reference,
)

MODEL = "gemma4:31b-cloud"
FORM_SHEET = "Форма заполнения предписания"
ROW_CONTENT = 18
ROW_NORMATIVE = 19
COL_VALUE = 2
COL_HINT = 3


SYSTEM_PROMPT = """Ты — ведущий инженер строительного контроля (СК). Проверяешь предписание по нормативной базе.

РЕЖИМ: ПРОФЕССИОНАЛЬНАЯ РЕДАКТУРА B18 + СВЕРКА B19 (БЕЗ ПРАВОК B19).
Перепиши только содержание замечания (B18). Нормативную ссылку (B19) НЕ переписывай и НЕ подменяй — только проверь по фрагменту из Техэксперт/интернет, насколько она обосновывает замечание.

На входе:
- B18 — текст инженера (исходник замечания).
- B19 — ссылка инженера на НД (только для проверки, не для правки).
- Фрагмент нормативки из онлайн-источника (если получен).

ЖЁСТКИЕ ПРАВИЛА B18 — ЧТО СОХРАНЯТЬ (объективные факты):
- объект, трубопровод, диаметры (Ø530×10 и т.п.);
- номера стыков/соединений (№ 2/2, 4, 8…);
- методы контроля (радиография, УЗК…), сроки («до начала гидроиспытаний»);
- требуемые действия (ремонт, повторный контроль);
- клеймо сварщика/бригады — только как идентификатор исполнителя, если это в исходнике.

ЖЁСТКИЕ ПРАВИЛА B18 — СТИЛЬ И ТОН:
1. Стиль предписания СК: безлично, нейтрально, по факту нарушения.
2. Убери оценочные суждения без цифр («большое количество брака», обвинения бригаде).
3. Не сжимай факты в пустые обобщения.
4. Не вставляй в B18 номера приказов, ГОСТ, пункты НД — только в B19.

ЖЁСТКИЕ ПРАВИЛА B19 — ТОЛЬКО ПРОВЕРКА, НЕ РЕДАКТИРОВАНИЕ:
1. ЗАПРЕЩЕНО менять текст B19, разворачивать название, подставлять другой документ или другие пункты.
2. Сверь B19 с фрагментом НД из источника: тот ли это документ, существуют ли указанные пункты.
3. Сопоставь требования из пунктов B19 с фактами в B18: обосновывает ли ссылка суть замечания.
4. Если ссылка слабая, документ не найден или пункт не по теме — опиши это в «СВЕРКА B19» и задай вопрос инженеру; B19 не переписывай.
5. Если фрагмент не получен — статус «не проверено онлайн», B19 не трогай.

АЛГОРИТМ:
1. Выпиши объективные факты из B18.
2. По фрагменту НД оцени документ и пункты из B19.
3. Сопоставь: соответствует ли нормативная ссылка содержанию замечания.
4. Перепиши B18 профессиональным языком СК.
5. При нехватке деталей — вопросы инженеру.

ФОРМАТ ОТВЕТА (строго):

## РЕЗЮМЕ ПРОВЕРКИ
- [✓ или ⚠] Сверка с нормативной базой: [Техэксперт / интернет / не получено]
- [✓ или ⚠] Содержание замечания (B18): [переработано / что изменено / недостаточно конкретики]
- [✓ или ⚠] Соответствие B19 замечанию: [соответствует / частично / не соответствует / не проверено]
- Решение: [переработка B18 / без изменений B18] — B19 не изменяется

## СВЕРКА B19 С ЗАМЕЧАНИЕМ
- Ссылка инженера (B19): [как указано, без правок]
- Документ в источнике: [что найдено в Техэксперт/интернет или «не найден»]
- Пункты B19: [перечень]
- Соответствие замечанию: [✓ / ⚠ / ✗ / не проверено]
- Обоснование: [связь фактов B18 с требованиями пунктов B19; 3–6 предложений]
- Замечания по ссылке: [если документ/пункт не подходят — что именно; иначе «замечаний нет»]

## ВОПРОСЫ ИНЖЕНЕРУ
(Если всё конкретно — одна строка: «Дополнительных вопросов нет.»)
1. [Вопрос] — [связь с B18/B19]

## ЧЕРНОВИК ПИСЬМА ИНЖЕНЕРУ
(2–5 предложений. Если вопросов нет — «Уточнения не требуются.»)

## ОТЧЁТ О ПРАВКАХ
### Содержание замечания (B18)
- Статус: [без изменений / переработано]
- Стиль и тон: [что убрано/перефразировано]
- Факты: [что сохранено]
- Причина правок: [кратко]

### Нормативный документ (B19)
- Статус: без изменений (только проверка)
- Сверка: [итог: соответствует / частично / не соответствует / не проверено]
- Вывод: [1–2 предложения]

## ИСПРАВЛЕННЫЕ ПОЛЯ
Содержание замечания:
[полный профессиональный текст B18]

(Блок «Нормативный документ» в ИСПРАВЛЕННЫЕ ПОЛЯ не выводи — B19 не меняется.)"""


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _open_form_sheet(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore[import-untyped]
        except ImportError as e:
            raise ValueError(
                "Формат .xls требует xlrd — сохраните как .xlsx или установите xlrd"
            ) from e
        book = xlrd.open_workbook(str(path))
        try:
            sheet = book.sheet_by_name(FORM_SHEET)
        except xlrd.XLRDError as e:
            raise ValueError(f"Лист «{FORM_SHEET}» не найден") from e
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = FORM_SHEET
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(r + 1, c + 1, value=sheet.cell_value(r, c))
        return wb, ws

    wb = load_workbook(path, data_only=True, read_only=False)
    if FORM_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Лист «{FORM_SHEET}» не найден")
    return wb, wb[FORM_SHEET]


def extract_form_fields(filepath: str | Path) -> dict[str, str]:
    """B18, B19 и подсказки из C18, C19."""
    path = Path(filepath)
    wb, ws = _open_form_sheet(path)
    try:
        fields = {
            "content": _cell_str(ws.cell(ROW_CONTENT, COL_VALUE).value),
            "normative": _cell_str(ws.cell(ROW_NORMATIVE, COL_VALUE).value),
            "content_hint": _cell_str(ws.cell(ROW_CONTENT, COL_HINT).value),
            "normative_hint": _cell_str(ws.cell(ROW_NORMATIVE, COL_HINT).value),
        }
    finally:
        wb.close()
    return fields


def _parse_corrected_fields(report_text: str) -> dict[str, str]:
    cleaned = re.sub(r"\*\*([^*]+)\*\*", r"\1", report_text)
    section = re.search(
        r"##\s*ИСПРАВЛЕННЫЕ\s*ПОЛЯ\s*(.*)$",
        cleaned,
        re.DOTALL | re.IGNORECASE,
    )
    body = section.group(1).strip() if section else ""

    content_m = re.search(
        r"Содержание\s+замечания\s*:\s*(.*?)(?=Нормативный\s+документ\s*:|$)",
        body,
        re.DOTALL | re.IGNORECASE,
    )
    norm_m = re.search(
        r"Нормативный\s+документ\s*:\s*(.*?)$",
        body,
        re.DOTALL | re.IGNORECASE,
    )
    out: dict[str, str] = {}
    if content_m:
        out["content"] = content_m.group(1).strip()
    if norm_m:
        out["normative"] = norm_m.group(1).strip()
    return out


def _number_in_text(num: str, text: str) -> bool:
    """Число или дробь (2/2) есть в тексте — не привязано к формату «№»."""
    if not num or not text:
        return False
    if num in text:
        return True
    if "/" in num:
        a, b = num.split("/", 1)
        return a in text and b in text
    return bool(re.search(rf"\b{re.escape(num)}\b", text))


def _missing_b18_facts(original: str, rewritten: str) -> list[str]:
    """
    Какие объективные факты из B18 инженера пропали в переработке.
    Сравнение по смыслу (числа, стыки, клеймо), а не по точной строке regex.
    """
    if not original or not rewritten:
        return []
    missing: list[str] = []

    for m in re.finditer(r"Ø\s*(\d+)\s*[xх×]\s*(\d+)", original, flags=re.IGNORECASE):
        d1, d2 = m.group(1), m.group(2)
        if not (_number_in_text(d1, rewritten) and _number_in_text(d2, rewritten)):
            missing.append(f"диаметр Ø{d1}×{d2}")

    seen_joints: set[str] = set()
    for m in re.finditer(
        r"(?:№№?|стык\w*|соединен\w*)\s*([\d\s.,/]+)",
        original,
        flags=re.IGNORECASE,
    ):
        for j in re.findall(r"\d+(?:/\d+)?", m.group(1)):
            if j in seen_joints:
                continue
            seen_joints.add(j)
            if not _number_in_text(j, rewritten):
                missing.append(f"стык/соединение №{j}")

    # Клеймо не обязательно: при снятии оценочных фраз о бригаде допустимо не повторять клеймо.

    if re.search(r"радиограф", original, flags=re.IGNORECASE):
        if not re.search(r"радиограф", rewritten, flags=re.IGNORECASE):
            missing.append("радиографический контроль")

    if re.search(r"неразрушающ", original, flags=re.IGNORECASE):
        if not re.search(r"неразрушающ", rewritten, flags=re.IGNORECASE):
            missing.append("неразрушающий контроль (НК)")

    if re.search(r"гидравлич|гидроиспыт", original, flags=re.IGNORECASE):
        if not re.search(r"гидравлич|гидроиспыт", rewritten, flags=re.IGNORECASE):
            missing.append("условие/срок до гидроиспытаний")

    return missing


def _usable_short_title(title: str, engineer_norm: str) -> bool:
    """Краткий title из выдачи пригоден для B19."""
    t = (title or "").strip()
    if len(t) < 12:
        return False
    ref = parse_normative_reference(engineer_norm)
    if ref.number and ref.number not in t:
        return False
    low = t.lower()
    if ref.search_query and low == ref.search_query.lower():
        return False
    if re.search(r"Об утверждении|Зарегистрировано в Минюсте", t, flags=re.IGNORECASE):
        return False
    return bool(
        re.search(
            r"(?:приказ|постановлен|гост|сп|снип|n\s*\d|№\s*\d)",
            t,
            flags=re.IGNORECASE,
        )
    )


def _b19_title_from_lookup(
    engineer_norm: str, normative_lookup: dict[str, Any]
) -> str:
    """Краткий title из источника (как в выдаче Техэксперт)."""
    ref = parse_normative_reference(engineer_norm)
    for raw in (
        normative_lookup.get("list_title"),
        normative_lookup.get("doc_title"),
    ):
        short = _short_doc_title(str(raw or ""), ref)
        if short and _usable_short_title(short, engineer_norm):
            return short.rstrip(" .,")
    built = _title_from_reference(ref)
    return built.rstrip(" .,") if built else ""


def _compose_b19(
    engineer_norm: str,
    normative_lookup: dict[str, Any],
) -> str:
    """B19: краткий title из источника + пункты инженера (модель не источник названия)."""
    pts = _normative_points(engineer_norm)
    pts_tail = ", ".join(f"п. {p}" for p in pts) if pts else ""

    if normative_lookup.get("ok"):
        base = _b19_title_from_lookup(engineer_norm, normative_lookup)
        if base:
            return f"{base}, {pts_tail}." if pts_tail else f"{base}."

    return engineer_norm.strip()


def _normative_source_info(lookup: dict[str, Any]) -> dict[str, str]:
    """Метаданные источника нормативки для UI и отчёта."""
    src = (lookup.get("source") or "unknown").lower()
    labels = {
        "techexpert": "Техэксперт",
        "internet": "Интернет (открытые источники)",
    }
    label = labels.get(src, src)
    if lookup.get("ok"):
        status = "получен"
    elif lookup.get("error"):
        status = "не получен"
    else:
        status = "не выполнялся"
    return {
        "source": src,
        "label": label,
        "status": status,
        "doc_title": lookup.get("doc_title") or "",
        "source_url": lookup.get("source_url") or "",
        "error": lookup.get("error") or "",
        "te_fallback_error": lookup.get("te_fallback_error") or "",
    }


def _parse_numbered_list(section: str) -> list[str]:
    if not section:
        return []
    items: list[str] = []
    for line in section.splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^\d+[\.)]\s*(.+)$", line)
        if m:
            items.append(m.group(1).strip())
        elif line.startswith("- ") and not items:
            items.append(line[2:].strip())
    return items


def _parse_engineer_questions(report_text: str) -> list[str]:
    section = _extract_report_section(report_text, "ВОПРОСЫ ИНЖЕНЕРУ")
    if not section:
        return []
    if re.search(r"дополнительных\s+вопросов\s+нет", section, re.I):
        return []
    return _parse_numbered_list(section)


def _parse_draft_letter(report_text: str) -> str:
    section = _extract_report_section(report_text, "ЧЕРНОВИК ПИСЬМА ИНЖЕНЕРУ")
    if not section:
        return ""
    if re.search(r"уточнения\s+не\s+требуются", section, re.I):
        return ""
    return section.strip()


def _parse_normative_assessment(report_text: str) -> str:
    """Текст сверки B19 с замечанием из отчёта модели."""
    section = _extract_report_section(report_text, "СВЕРКА B19 С ЗАМЕЧАНИЕМ")
    if section:
        return section.strip()
    block = _extract_report_section(report_text, "ОТЧЁТ О ПРАВКАХ")
    if not block:
        return ""
    m = re.search(
        r"###\s*Нормативный\s+документ\s*\(B19\)\s*(.*?)(?=###|\Z)",
        block,
        re.DOTALL | re.IGNORECASE,
    )
    return m.group(1).strip() if m else ""


_SUBJECTIVE_RE = re.compile(
    r"больш(ое|ая|ой)\s+количеств|"
    r"систематическ|"
    r"постоянн\w*\s+(?:допуск|наруш)|"
    r"работает\s+плохо|"
    r"допускает\s+брак|"
    r"много\s+брака|"
    r"недобросовест",
    re.I,
)


_DEFECT_TYPE_RE = re.compile(
    r"трещин|пор[ыа]|шлак|подрез|несплавлен|включен|вклинен|"
    r"смещен|непровар|расслоен|карман|выгоран|окисн|"
    r"подрезан|вогнут|выпукл|отклонен",
    re.I,
)


def _rule_based_questions(
    fields: dict[str, str], normative_lookup: dict[str, Any]
) -> list[str]:
    """Эвристики, если модель не задала вопросы по размытым формулировкам."""
    content = fields.get("content") or ""
    if not content:
        return []
    questions: list[str] = []
    pts = _normative_points(fields.get("normative") or "")
    pts_hint = f" (п. {', '.join(pts)})" if pts else ""

    if re.search(r"дефект", content, re.I) and not _DEFECT_TYPE_RE.search(content):
        questions.append(
            "Какие именно недопустимые дефекты выявлены (тип, размер, расположение по каждому стыку/соединению)?"
            f"{pts_hint} — в B18 указано только «дефекты» без конкретики."
        )

    if re.search(r"недопустим", content, re.I) and not _DEFECT_TYPE_RE.search(content):
        if not questions:
            questions.append(
                "Чем обоснована оценка «недопустимо» — какие параметры дефекта превышают норму по НД"
                f"{pts_hint}?"
            )

    if _SUBJECTIVE_RE.search(content):
        questions.append(
            "В исходнике есть оценочные формулировки («большое количество брака», "
            "«допускает брак» и т.п.) без цифр. Подтвердите количество дефектных стыков "
            "или согласны убрать оценочный тон из предписания?"
        )

    if re.search(r"брак", content, re.I) and not _SUBJECTIVE_RE.search(content) and not re.search(
        r"количеств|процент|стык|соединен|шов", content, re.I
    ):
        questions.append(
            "В чём проявляется «брак»: перечень дефектных стыков, виды дефектов, статистика?"
        )

    if re.search(r"ремонт", content, re.I) and not re.search(
        r"выруб|шлиф|перевар|зачист|способ", content, re.I
    ):
        questions.append(
            "Какой способ ремонта сварных соединений предусмотрен (вырубка, шлифовка, переварка и т.п.)?"
        )

    if normative_lookup.get("ok") and pts and not questions:
        excerpt = (normative_lookup.get("excerpt") or "")[:800]
        if excerpt and re.search(r"дефект|контрол|качеств", excerpt, re.I):
            if not re.search(r"акт|заключен|протокол|снимок|пленк", content, re.I):
                questions.append(
                    f"Приложены ли акты/заключения НК по стыкам, на которые ссылается предписание{pts_hint}?"
                )

    return questions


def _merge_questions(model_q: list[str], rule_q: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for q in model_q + rule_q:
        key = q.lower()[:80]
        if key in seen:
            continue
        seen.add(key)
        out.append(q)
    return out


def _normative_points(text: str) -> list[str]:
    if not text:
        return []
    return parse_normative_reference(text).points


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _text_changed(before: str, after: str) -> bool:
    return _normalize_text(before) != _normalize_text(after)


def _extract_report_section(report_text: str, heading: str) -> str:
    pattern = rf"##\s*{re.escape(heading)}\s*(.*?)(?=##\s*|\Z)"
    m = re.search(pattern, report_text, re.DOTALL | re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _guard_engineer_text(
    fields: dict[str, str],
    corrected: dict[str, str],
    normative_lookup: dict[str, Any] | None = None,
) -> tuple[dict[str, str], list[dict[str, str]]]:
    """
    B18: не терять объективные факты при переработке.
    B19: не изменять — только сверка в отчёте.
    """
    out = dict(corrected)
    events: list[dict[str, str]] = []
    orig_content = fields.get("content") or ""
    orig_norm = fields.get("normative") or ""
    new_content = out.get("content") or ""
    if orig_content and new_content and _text_changed(orig_content, new_content):
        missing = _missing_b18_facts(orig_content, new_content)
        if missing:
            lost_str = "; ".join(missing)
            print(f"[PRESCRIPTION_CHECK] B18 guard: revert — {lost_str}")
            events.append(
                {
                    "field": "content",
                    "level": "warn",
                    "code": "guard_content_revert",
                    "message": (
                        "Переработка B18 не записана в файл: в тексте модели не найдены "
                        f"факты из исходника ({lost_str}). См. блок «Предложение модели»."
                    ),
                    "model_preview": new_content,
                    "lost_facts": lost_str,
                }
            )
            out["content"] = orig_content

    if orig_norm:
        out["normative"] = orig_norm
        model_norm = corrected.get("normative") or ""
        if model_norm and _text_changed(orig_norm, model_norm):
            events.append(
                {
                    "field": "normative",
                    "level": "info",
                    "code": "normative_preserved",
                    "message": (
                        "B19 не изменён: сохранена ссылка инженера. "
                        "Сверка с замечанием — в блоке «Сверка B19»."
                    ),
                    "model_preview": model_norm,
                }
            )

    return out, events


def _report_has_issues(report_text: str) -> bool:
    resume = _extract_report_section(report_text, "РЕЗЮМЕ ПРОВЕРКИ")
    block = resume or report_text
    return "⚠" in block


def _log_prescription_fields(
    filename: str,
    fields: dict[str, str],
    model_corrected: dict[str, str],
    final_corrected: dict[str, str],
    normative_assessment: str = "",
) -> None:
    """Пишет в консоль B18 (исходник/агент/файл) и B19 (исходник/сверка)."""
    engineer_b18 = (fields.get("content") or "").strip()
    agent_b18 = (model_corrected.get("content") or "").strip()
    final_b18 = (final_corrected.get("content") or "").strip()
    if engineer_b18 or agent_b18 or final_b18:
        print(f"[PRESCRIPTION_CHECK] --- {filename} B18 инженер (замечание) ---")
        print(engineer_b18 or "(пусто)")
        print(f"[PRESCRIPTION_CHECK] --- {filename} B18 агент (замечание) ---")
        print(agent_b18 or "(без переработки)")
        if final_b18 and final_b18 != agent_b18 and final_b18 != engineer_b18:
            print(f"[PRESCRIPTION_CHECK] --- {filename} B18 в файл (замечание) ---")
            print(final_b18)

    engineer_b19 = (fields.get("normative") or "").strip()
    if engineer_b19:
        print(f"[PRESCRIPTION_CHECK] --- {filename} B19 инженер (нормативка) ---")
        print(engineer_b19)
        print(f"[PRESCRIPTION_CHECK] --- {filename} B19 сверка с замечанием ---")
        print(normative_assessment.strip() or "(нет блока сверки в отчёте модели)")
        print(f"[PRESCRIPTION_CHECK] --- {filename} B19 в файл ---")
        print(engineer_b19 + " (без изменений)")


def _field_compare_lines(
    field_label: str,
    orig: str,
    agent: str,
    final: str,
) -> list[str] | None:
    orig = (orig or "").strip()
    agent = (agent or "").strip()
    final = (final or "").strip()
    if not orig and not agent and not final:
        return None
    lines = [
        f"### Исходник инженера ({field_label})",
        orig or "(пусто)",
        "",
        f"### Переработка агента ({field_label})",
        agent or "(без переработки)",
    ]
    if final and final not in {orig, agent}:
        lines.extend(["", f"### Записано в файл ({field_label})", final])
    return lines


def _build_review_display(
    report_text: str,
    fields: dict[str, str],
    model_corrected: dict[str, str],
    final_corrected: dict[str, str],
    guard_events: list[dict[str, str]],
    normative_lookup: dict[str, Any],
    engineer_questions: list[str],
    draft_letter: str,
    normative_assessment: str = "",
) -> str:
    """Текст отчёта для панели UI."""
    parts: list[str] = []
    src = _normative_source_info(normative_lookup)

    source_lines = [f"**Источник нормативки:** {src['label']} — {src['status']}"]
    if src["doc_title"]:
        source_lines.append(f"**Документ:** {src['doc_title']}")
    if src["source_url"]:
        source_lines.append(f"**URL:** {src['source_url']}")
    if src.get("te_fallback_error"):
        source_lines.append(
            f"**Техэксперт (не использован):** {src['te_fallback_error']}"
        )
    if src["error"] and not normative_lookup.get("ok"):
        source_lines.append(f"**Ошибка поиска:** {src['error']}")
    parts.append("## ИСТОЧНИК НОРМАТИВКИ\n" + "\n".join(f"- {l}" for l in source_lines))

    if engineer_questions:
        q_lines = "\n".join(f"{i}. {q}" for i, q in enumerate(engineer_questions, 1))
        parts.append("## ВОПРОСЫ ИНЖЕНЕРУ\n" + q_lines)
    else:
        parts.append("## ВОПРОСЫ ИНЖЕНЕРУ\nДополнительных вопросов нет — предписание достаточно конкретно.")

    if draft_letter:
        parts.append("## ЧЕРНОВИК ПИСЬМА ИНЖЕНЕРУ\n" + draft_letter)

    orig_content = (fields.get("content") or "").strip()
    agent_content = (model_corrected.get("content") or "").strip()
    final_content = (final_corrected.get("content") or "").strip()
    b18_lines = _field_compare_lines("B18", orig_content, agent_content, final_content)
    if b18_lines:
        parts.append("## СОДЕРЖАНИЕ ЗАМЕЧАНИЯ (B18)\n" + "\n".join(b18_lines))

    orig_norm = (fields.get("normative") or "").strip()
    if orig_norm:
        b19_lines = [
            "### Исходник инженера (B19)",
            orig_norm,
            "",
            "### Сверка с замечанием (B19 не изменяется)",
            normative_assessment.strip() or "(модель не заполнила блок сверки)",
        ]
        model_norm = (model_corrected.get("normative") or "").strip()
        if model_norm and _text_changed(orig_norm, model_norm):
            b19_lines.extend(
                [
                    "",
                    "### Ошибочное предложение модели по B19 (не применено)",
                    model_norm,
                ]
            )
        parts.append("## НОРМАТИВНЫЙ ДОКУМЕНТ (B19)\n" + "\n".join(b19_lines))

    resume = _extract_report_section(report_text, "РЕЗЮМЕ ПРОВЕРКИ")
    changes = _extract_report_section(report_text, "ОТЧЁТ О ПРАВКАХ")
    if resume:
        parts.append("## РЕЗЮМЕ ПРОВЕРКИ\n" + resume)
    if changes:
        parts.append("## ОТЧЁТ О ПРАВКАХ\n" + changes)

    model_blocks: list[str] = []
    if model_corrected.get("content") and _text_changed(
        fields.get("content", ""), model_corrected.get("content", "")
    ):
        model_blocks.append(
            "### B18 — как предложила модель\n" + model_corrected["content"]
        )
    if model_corrected.get("normative") and _text_changed(
        fields.get("normative", ""), model_corrected.get("normative", "")
    ):
        model_blocks.append(
            "### B19 — ошибочное предложение модели (не записано в файл)\n"
            + model_corrected["normative"]
        )
    if model_blocks and _text_changed(
        model_corrected.get("content", ""), final_corrected.get("content", "")
    ):
        parts.append("## ПРЕДЛОЖЕНИЕ МОДЕЛИ\n" + "\n\n".join(model_blocks))

    applied: list[str] = []
    b18_guard = next((e for e in guard_events if e.get("code") == "guard_content_revert"), None)
    if _text_changed(fields.get("content", ""), final_corrected.get("content", "")):
        applied.append("B18 — в файл записана переработка модели")
    elif b18_guard:
        applied.append(
            "B18 — переработка модели не записана: "
            + (b18_guard.get("lost_facts") or "потеряны факты из исходника")
        )
    elif _text_changed(fields.get("content", ""), model_corrected.get("content", "")):
        applied.append("B18 — модель предлагала правки, в файл записан исходник")
    else:
        applied.append("B18 — без изменений")

    applied.append("B19 — без изменений (только сверка в отчёте)")

    parts.append("## ИТОГ ЗАПИСИ В ФАЙЛ\n" + "\n".join(f"- {line}" for line in applied))

    if guard_events:
        parts.append(
            "## ЗАМЕЧАНИЯ СИСТЕМЫ\n"
            + "\n".join(f"- ⚠ {e['message']}" for e in guard_events)
        )

    return "\n\n".join(parts).strip()


def _collect_issues(
    report_text: str,
    fields: dict[str, str],
    model_corrected: dict[str, str],
    final_corrected: dict[str, str],
    guard_events: list[dict[str, str]],
    normative_lookup: dict[str, Any],
    engineer_questions: list[str],
    normative_assessment: str = "",
) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    src = _normative_source_info(normative_lookup)

    if fields.get("normative"):
        if normative_lookup.get("ok"):
            msg = f"Нормативка: {src['label']}"
            if src["doc_title"]:
                msg += f" — {src['doc_title'][:80]}"
            if src.get("te_fallback_error"):
                msg += f" (Техэксперт: {src['te_fallback_error'][:100]})"
            issues.append(
                {
                    "level": "info",
                    "code": "normative_source",
                    "message": msg,
                }
            )
        else:
            issues.append(
                {
                    "level": "warn",
                    "code": "normative_lookup",
                    "message": normative_lookup.get("error")
                    or "Не удалось получить текст нормативного документа для сверки",
                }
            )

    if engineer_questions:
        issues.append(
            {
                "level": "warn",
                "code": "clarification_needed",
                "message": (
                    f"Нужно уточнение у инженера: {len(engineer_questions)} "
                    f"вопрос(ов) — см. блок «Вопросы инженеру»"
                ),
            }
        )

    for ev in guard_events:
        issues.append(
            {
                "level": ev.get("level") or "warn",
                "code": ev.get("code") or "guard",
                "message": ev.get("message") or "Система отклонила правку модели",
            }
        )

    if fields.get("normative") and normative_assessment:
        if re.search(r"не\s+соответств|✗|не\s+подход", normative_assessment, re.I):
            issues.append(
                {
                    "level": "warn",
                    "code": "normative_mismatch",
                    "message": (
                        "Нормативная ссылка (B19) может не обосновывать замечание — "
                        "см. «Сверка B19»"
                    ),
                }
            )
        elif re.search(r"частичн|⚠", normative_assessment, re.I):
            issues.append(
                {
                    "level": "warn",
                    "code": "normative_partial",
                    "message": "Соответствие B19 замечанию частичное — см. «Сверка B19»",
                }
            )

    if _text_changed(fields.get("content", ""), final_corrected.get("content", "")):
        reason = _extract_field_reason(report_text, "B18") or "см. отчёт о правках"
        issues.append(
            {
                "level": "warn",
                "code": "content_changed",
                "message": f"Изменено содержание замечания (B18): {reason}",
            }
        )

    if _report_has_issues(report_text) and not issues:
        issues.append(
            {
                "level": "warn",
                "code": "model_flags",
                "message": "Модель отметила замечания в резюме проверки",
            }
        )

    return issues


def _extract_field_reason(report_text: str, field: str) -> str:
    block = _extract_report_section(report_text, "ОТЧЁТ О ПРАВКАХ")
    if not block:
        return ""
    if field.upper() == "B18":
        section = re.search(
            r"###\s*Содержание\s+замечания.*?-\s*Причина:\s*(.+?)(?=###|\Z)",
            block,
            re.DOTALL | re.IGNORECASE,
        )
    else:
        section = re.search(
            r"###\s*Нормативный\s+документ.*?-\s*Причина:\s*(.+?)(?=###|\Z)",
            block,
            re.DOTALL | re.IGNORECASE,
        )
    if not section:
        return ""
    reason = section.group(1).strip()
    reason = re.sub(r"\s*-\s*Статус:.*", "", reason, flags=re.DOTALL)
    return reason.split("\n")[0].strip()[:300]


def _format_normative_lookup(lookup: dict[str, Any]) -> str:
    ref = lookup.get("reference") or {}
    src = lookup.get("source") or "techexpert"
    src_label = {"techexpert": "Техэксперт", "internet": "интернет"}.get(src, src)
    lines = [
        f"Источник нормативки: {src_label}",
        f"Поисковый запрос: {ref.get('search_query') or '—'}",
    ]
    if lookup.get("doc_title"):
        lines.append(f"Найденный документ: {lookup['doc_title']}")
    if lookup.get("source_url"):
        lines.append(f"URL: {lookup['source_url']}")
    if lookup.get("te_fallback_error"):
        lines.append(f"Техэксперт недоступен: {lookup['te_fallback_error']}")
    if lookup.get("ok") and lookup.get("excerpt"):
        lines.append("")
        lines.append(f"--- Фрагмент из {src_label} ---")
        lines.append(str(lookup["excerpt"]))
        lines.append("--- конец фрагмента ---")
    elif lookup.get("error"):
        lines.append(f"Ошибка: {lookup['error']}")
    return "\n".join(lines)


def check_prescription(filepath: str | Path) -> dict:
    """Проверить предписание: Техэксперт + Ollama."""
    path = Path(filepath)
    filename = path.name

    try:
        fields = extract_form_fields(path)
    except Exception as e:
        return {
            "ok": False,
            "report": f"Ошибка чтения формы: {e}",
            "_source_file": filename,
            "issues": [{"level": "error", "message": str(e)}],
        }

    if not fields["content"] and not fields["normative"]:
        return {
            "ok": False,
            "report": "Ячейки B18 и B19 пустые — нечего проверять",
            "_source_file": filename,
            "issues": [{"level": "error", "message": "Пустые B18 и B19"}],
        }

    normative_lookup: dict[str, Any] = {"ok": False, "error": "B19 пуст — поиск не выполнялся"}
    if fields["normative"]:
        try:
            normative_lookup = lookup_normative(fields["normative"])
        except Exception as e:
            print(f"[PRESCRIPTION_CHECK] techexpert error: {e}")
            normative_lookup = {"ok": False, "error": str(e)}

    user_prompt = f"""Проверь предписание (лист «{FORM_SHEET}») по нормативной базе.

Подсказки шаблона:
- Содержание замечания: {fields["content_hint"] or "—"}
- Нормативный документ: {fields["normative_hint"] or "—"}

--- Содержание замечания (B18) ---
{fields["content"] or "(пусто)"}
--- конец ---

--- Нормативный документ (B19) ---
{fields["normative"] or "(пусто)"}
--- конец ---

{_format_normative_lookup(normative_lookup)}

Пункты НД, указанные инженером в B19: {", ".join(_normative_points(fields["normative"])) or "—"}

ВАЖНО:
- B18: перепиши профессиональным языком СК — конкретно, без воды; сохрани объективные факты.
- B19: НЕ переписывай. Сверь ссылку инженера с фрагментом НД и с фактами B18; результат — в «СВЕРКА B19 С ЗАМЕЧАНИЕМ».
- В «ИСПРАВЛЕННЫЕ ПОЛЯ» укажи только B18. Блок B19 не выводи.
- Заполни «ВОПРОСЫ ИНЖЕНЕРУ», «ЧЕРНОВИК ПИСЬМА» и «ОТЧЁТ О ПРАВКАХ» (в т.ч. «Стиль и тон» для B18).

Ответ строго в формате из инструкции."""

    try:
        import ollama

        response = ollama.chat(
            model=MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            stream=False,
        )
        report_text = response.get("message", {}).get("content", "").strip()
    except Exception as e:
        print(f"[PRESCRIPTION_CHECK] ollama error: {e}")
        return {
            "ok": False,
            "report": f"Ошибка проверки: {e}",
            "_source_file": filename,
            "normative_lookup": normative_lookup,
            "issues": [{"level": "error", "message": str(e)}],
        }

    if not report_text:
        return {
            "ok": False,
            "report": "Ошибка: пустой ответ модели",
            "_source_file": filename,
            "normative_lookup": normative_lookup,
            "issues": [{"level": "error", "message": "пустой ответ модели"}],
        }

    model_corrected = _parse_corrected_fields(report_text)
    corrected, guard_events = _guard_engineer_text(
        fields, model_corrected, normative_lookup
    )
    model_questions = _parse_engineer_questions(report_text)
    rule_questions = _rule_based_questions(fields, normative_lookup)
    engineer_questions = _merge_questions(model_questions, rule_questions)
    draft_letter = _parse_draft_letter(report_text)
    normative_assessment = _parse_normative_assessment(report_text)
    normative_source = _normative_source_info(normative_lookup)
    review_display = _build_review_display(
        report_text,
        fields,
        model_corrected,
        corrected,
        guard_events,
        normative_lookup,
        engineer_questions,
        draft_letter,
        normative_assessment,
    )
    issues = _collect_issues(
        report_text,
        fields,
        model_corrected,
        corrected,
        guard_events,
        normative_lookup,
        engineer_questions,
        normative_assessment,
    )
    has_errors = any(i.get("level") == "error" for i in issues)
    has_warnings = any(i.get("level") == "warn" for i in issues)

    result = {
        "ok": not has_errors and not has_warnings,
        "report": report_text,
        "review_display": review_display,
        "_source_file": filename,
        "fields": fields,
        "corrected": corrected,
        "model_corrected": model_corrected,
        "guard_events": guard_events,
        "normative_lookup": normative_lookup,
        "normative_source": normative_source,
        "engineer_questions": engineer_questions,
        "draft_letter": draft_letter,
        "normative_assessment": normative_assessment,
        "issues": issues,
        "has_errors": has_errors,
        "has_warnings": has_warnings,
    }
    print(
        f"[PRESCRIPTION_CHECK] "
        f"{'OK' if result['ok'] else ('ERR' if has_errors else 'WARN')}: {filename} "
        f"(te={'ok' if normative_lookup.get('ok') else 'fail'}, "
        f"issues={len(issues)})"
    )
    _log_prescription_fields(
        filename, fields, model_corrected, corrected, normative_assessment
    )
    return result


def write_checked_copy(src: str | Path, dest: str | Path, check_result: dict) -> None:
    """Копия файла с исправленным B18 (B19 не меняется)."""
    src_path = Path(src)
    dest_path = Path(dest)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_path, dest_path)

    corrected = check_result.get("corrected") or {}
    if not corrected.get("content"):
        return
    if dest_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return

    wb = load_workbook(dest_path)
    if FORM_SHEET not in wb.sheetnames:
        wb.close()
        return
    ws = wb[FORM_SHEET]
    ws.cell(ROW_CONTENT, COL_VALUE, value=corrected["content"])
    wb.save(dest_path)
    wb.close()
