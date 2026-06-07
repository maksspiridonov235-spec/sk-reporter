"""Проверка файлов предписаний (.xlsx / .xls).

Логика ориентирована на типовую таблицу: шапка с колонками про номер, дату,
объект, нарушение, срок устранения. Правила расширяются по мере появления
эталонного шаблона в data/prescriptions/.
"""

from __future__ import annotations

import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
_HEADER_SCAN_ROWS = 20
_ISSUE_COL_TITLE = "Проверка SK"

# Группы: хотя бы одна колонка из группы должна быть найдена в шапке.
_REQUIRED_HEADER_GROUPS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("номер", ("№", "номер", "no", "n/p", "п/п")),
    ("дата", ("дата", "date")),
    ("нарушение", ("наруш", "содержан", "описан", "замечан", "дефект", "выявлен")),
)

_RECOMMENDED_HEADER_GROUPS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("объект", ("объект", "наименован", "строитель")),
    ("срок", ("срок", "устранен", "deadline")),
    ("ответственный", ("ответств", "исполнит", "подрядчик")),
)


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _row_has_content(row: tuple[Any, ...]) -> bool:
    return any(_cell_str(v) for v in row)


def _match_group(header: str, keywords: tuple[str, ...]) -> bool:
    return any(kw in header for kw in keywords)


def _find_header_row(ws: Worksheet) -> tuple[int, list[str]] | None:
    best: tuple[int, list[str], int] | None = None
    for r in range(1, min(ws.max_row, _HEADER_SCAN_ROWS) + 1):
        values = [_cell_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not _row_has_content(values):
            continue
        norms = [_norm_header(v) for v in values]
        score = sum(
            1
            for _, kws in _REQUIRED_HEADER_GROUPS + _RECOMMENDED_HEADER_GROUPS
            if any(_match_group(h, kws) for h in norms if h)
        )
        if score >= 2 and (best is None or score > best[2]):
            best = (r, values, score)
    if best is None:
        return None
    return best[0], best[1]


def _map_columns(headers: list[str]) -> dict[str, int | None]:
    norms = [_norm_header(h) for h in headers]
    mapping: dict[str, int | None] = {}

    def pick(groups: tuple[tuple[str, tuple[str, ...]], ...]) -> None:
        for name, kws in groups:
            col_idx = None
            for i, h in enumerate(norms):
                if h and _match_group(h, kws):
                    col_idx = i
                    break
            mapping[name] = col_idx

    pick(_REQUIRED_HEADER_GROUPS)
    pick(_RECOMMENDED_HEADER_GROUPS)
    return mapping


def _parse_date(value: Any) -> date | None:
    if value is None or _cell_str(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _load_sheet(path: Path) -> tuple[Workbook, Worksheet, str]:
    suffix = path.suffix.lower()
    if suffix not in _EXCEL_SUFFIXES:
        raise ValueError(f"Неподдерживаемый формат: {suffix}")

    if suffix == ".xls":
        try:
            import xlrd  # type: ignore[import-untyped]
        except ImportError as e:
            raise ValueError(
                "Формат .xls требует пакет xlrd — сохраните файл как .xlsx или установите xlrd"
            ) from e
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        from openpyxl import Workbook as OWB

        wb = OWB()
        ws = wb.active
        ws.title = sheet.name or "Sheet1"
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(r + 1, c + 1, value=sheet.cell_value(r, c))
        return wb, ws, ws.title

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    return wb, ws, ws.title


def _issue(level: str, code: str, message: str, *, row: int | None = None, col: str | None = None) -> dict:
    return {"level": level, "code": code, "message": message, "row": row, "col": col}


def check_prescription_file(path: str | Path) -> dict:
    """Проверить один Excel-файл предписаний. Возвращает структуру для UI/API."""
    file_path = Path(path)
    issues: list[dict] = []
    summary: dict[str, Any] = {
        "filename": file_path.name,
        "sheet": None,
        "header_row": None,
        "rows_checked": 0,
        "rows_with_issues": 0,
    }

    if not file_path.is_file():
        issues.append(_issue("error", "file_missing", "Файл не найден"))
        return {"ok": False, "issues": issues, "summary": summary, "report": _format_report(issues)}

    try:
        wb, ws, sheet_name = _load_sheet(file_path)
    except Exception as e:
        issues.append(_issue("error", "open_failed", f"Не удалось открыть Excel: {e}"))
        return {"ok": False, "issues": issues, "summary": summary, "report": _format_report(issues)}

    summary["sheet"] = sheet_name

    if ws.max_row < 2 or ws.max_column < 2:
        issues.append(_issue("error", "empty_sheet", "Лист пустой или без данных"))
        wb.close()
        return {"ok": False, "issues": issues, "summary": summary, "report": _format_report(issues)}

    header_info = _find_header_row(ws)
    if header_info is None:
        issues.append(
            _issue(
                "error",
                "header_not_found",
                "Не найдена строка заголовков (ожидаются колонки: номер, дата, нарушение)",
            )
        )
        wb.close()
        return {"ok": False, "issues": issues, "summary": summary, "report": _format_report(issues)}

    header_row, headers = header_info
    summary["header_row"] = header_row
    cols = _map_columns(headers)

    for group_name, _ in _REQUIRED_HEADER_GROUPS:
        if cols.get(group_name) is None:
            issues.append(
                _issue(
                    "error",
                    "missing_column",
                    f"Не найдена обязательная колонка: «{group_name}»",
                )
            )

    for group_name, _ in _RECOMMENDED_HEADER_GROUPS:
        if cols.get(group_name) is None:
            issues.append(
                _issue(
                    "warn",
                    "recommended_column",
                    f"Рекомендуется колонка: «{group_name}»",
                )
            )

    if any(i["level"] == "error" and i["code"] == "missing_column" for i in issues):
        wb.close()
        return {"ok": False, "issues": issues, "summary": summary, "report": _format_report(issues)}

    data_start = header_row + 1
    rows_with_issues = 0
    row_issue_text: dict[int, list[str]] = {}

    for r in range(data_start, ws.max_row + 1):
        row_vals = tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
        if not _row_has_content(row_vals):
            continue

        summary["rows_checked"] += 1
        row_msgs: list[str] = []

        def col_val(name: str) -> Any:
            idx = cols.get(name)
            if idx is None:
                return None
            return ws.cell(r, idx + 1).value

        num_val = _cell_str(col_val("номер"))
        date_val = col_val("дата")
        violation_val = _cell_str(col_val("нарушение"))
        deadline_val = col_val("срок")
        obj_val = _cell_str(col_val("объект"))

        if not num_val:
            row_msgs.append("нет номера")
        if not violation_val:
            row_msgs.append("пустое описание нарушения")

        parsed_date = _parse_date(date_val)
        if _cell_str(date_val) and parsed_date is None:
            row_msgs.append(f"некорректная дата: «{_cell_str(date_val)}»")

        parsed_deadline = _parse_date(deadline_val)
        if _cell_str(deadline_val) and parsed_deadline is None:
            row_msgs.append(f"некорректный срок: «{_cell_str(deadline_val)}»")

        if parsed_date and parsed_deadline and parsed_deadline < parsed_date:
            row_msgs.append("срок устранения раньше даты выдачи")

        if cols.get("объект") is not None and not obj_val:
            row_msgs.append("не указан объект")

        if row_msgs:
            rows_with_issues += 1
            msg = "; ".join(row_msgs)
            row_issue_text[r] = row_msgs
            issues.append(
                _issue(
                    "warn" if violation_val else "error",
                    "row_invalid",
                    f"Строка {r}: {msg}",
                    row=r,
                )
            )

    summary["rows_with_issues"] = rows_with_issues

    if summary["rows_checked"] == 0:
        issues.append(_issue("error", "no_data_rows", "Нет строк с данными под заголовком"))

    wb.close()

    has_errors = any(i["level"] == "error" for i in issues)
    ok = not has_errors and summary["rows_checked"] > 0

    return {
        "ok": ok,
        "issues": issues,
        "summary": summary,
        "row_markers": {str(k): v for k, v in row_issue_text.items()},
        "report": _format_report(issues, summary),
    }


def _format_report(issues: list[dict], summary: dict | None = None) -> str:
    lines: list[str] = []
    if summary:
        lines.append(
            f"Лист: {summary.get('sheet')}, строка заголовка: {summary.get('header_row')}, "
            f"проверено строк: {summary.get('rows_checked', 0)}"
        )
    if not issues:
        lines.append("Замечаний нет.")
        return "\n".join(lines)
    for item in issues:
        prefix = {"error": "ОШИБКА", "warn": "Замечание"}.get(item["level"], item["level"])
        lines.append(f"[{prefix}] {item['message']}")
    return "\n".join(lines)


def write_checked_copy(src: str | Path, dest: str | Path, check_result: dict) -> None:
    """Копия файла с колонкой «Проверка SK» — текст замечаний по строкам."""
    src_path = Path(src)
    dest_path = Path(dest)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_path, dest_path)

    wb = load_workbook(dest_path)
    ws = wb.active
    header_row = check_result.get("summary", {}).get("header_row")
    if not header_row:
        wb.save(dest_path)
        wb.close()
        return

    issue_col = ws.max_column + 1
    ws.cell(header_row, issue_col, value=_ISSUE_COL_TITLE)

    markers = check_result.get("row_markers") or {}
    for row_s, msgs in markers.items():
        row = int(row_s)
        ws.cell(row, issue_col, value="; ".join(msgs))

    for r in range(header_row + 1, ws.max_row + 1):
        if str(r) in markers:
            continue
        if _row_has_content(tuple(ws.cell(r, c).value for c in range(1, issue_col))):
            ws.cell(r, issue_col, value="OK")

    wb.save(dest_path)
    wb.close()
