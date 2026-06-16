"""Парсинг справочника персонала из Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sk_reporter.personnel_store import person_id_from_fio


def parse_personnel_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Пустой лист в справочнике персонала")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    people: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        if not any(rec.values()):
            continue
        fio = " ".join(str(rec.get("ФИО") or "").split())
        if not fio:
            continue
        pid = str(rec.get("id") or person_id_from_fio(fio))
        if pid in seen:
            pid = f"{pid}-{len(seen)}"
        seen.add(pid)
        people.append(
            {
                "id": pid,
                "fio": fio,
                "phone": str(rec.get("Телефон") or "").strip(),
                "position": str(rec.get("Должность") or "").strip(),
                "control_mode": str(rec.get("Режим контроля") or "").strip(),
            }
        )
    return people
