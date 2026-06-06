#!/usr/bin/env python3
"""Собрать кэши данных инженера: vor.json, tk/manifest.yaml, personnel.yaml."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import yaml
from openpyxl import load_workbook

from sk_reporter.engineer.tk_catalog import write_manifest
from sk_reporter.personnel_store import person_id_from_fio
from sk_reporter.paths import personnel_dir, projects_dir


def export_personnel() -> Path:
    xlsx = personnel_dir() / "Справочник персонала.xlsx"
    if not xlsx.is_file():
        raise FileNotFoundError(xlsx)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Пустой лист в справочнике персонала")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    people = []
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        if not any(rec.values()):
            continue
        fio = str(rec.get("ФИО") or "").strip()
        if fio:
            rec["id"] = person_id_from_fio(fio)
        people.append(rec)
    out = personnel_dir() / "personnel.yaml"
    out.write_text(yaml.safe_dump({"people": people}, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return out


def build_vor_caches(project_ids: list[str] | None = None) -> list[Path]:
    from sk_reporter.engineer.vor_legacy import write_project_vor_cache

    built: list[Path] = []
    for proj in sorted(projects_dir().iterdir()):
        if not proj.is_dir() or proj.name.startswith("."):
            continue
        if project_ids and proj.name not in project_ids:
            continue
        meta_path = proj / "project.yaml"
        if not meta_path.is_file():
            continue
        meta = yaml.safe_load(meta_path.read_text(encoding="utf-8")) or {}
        has_vor = bool(meta.get("vor_docx")) or bool(meta.get("vor_doc"))
        if not has_vor:
            continue
        try:
            built.append(write_project_vor_cache(proj))
        except FileNotFoundError:
            continue
    return built


def main() -> None:
    parser = argparse.ArgumentParser(description="Build engineer data caches")
    parser.add_argument("--personnel", action="store_true", help="Export personnel.yaml")
    parser.add_argument("--tk", action="store_true", help="Write data/tk/manifest.yaml")
    parser.add_argument("--vor", action="store_true", help="Parse VOR docx → vor.json")
    parser.add_argument("--luvr", action="store_true", help="Export luvr.yaml from xlsx")
    parser.add_argument("--all", action="store_true", help="All of the above")
    parser.add_argument("--project", action="append", dest="projects", help="Project id filter")
    args = parser.parse_args()

    if args.all or not any((args.personnel, args.tk, args.vor, args.luvr)):
        args.personnel = args.tk = args.vor = args.luvr = True

    results: dict[str, str] = {}
    if args.personnel:
        results["personnel"] = str(export_personnel())
    if args.luvr:
        from sk_reporter.luvr_store import export_luvr

        results["luvr"] = str(export_luvr())
    if args.tk:
        results["tk_manifest"] = str(write_manifest())
    if args.vor:
        for p in build_vor_caches(args.projects):
            results.setdefault("vor", []).append(str(p))  # type: ignore[union-attr]

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
